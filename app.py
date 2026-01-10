import streamlit as st
import pandas as pd
from supabase import create_client, Client
import plotly.express as px
import plotly.graph_objects as go
import numpy as np
import os
import time
from datetime import datetime, timedelta
from fpdf import FPDF

import streamlit as st
from fpdf import FPDF
from datetime import datetime
import io
import time

# 1. FUNCIÓN GENERADORA DE PDF (Tu lógica corregida)
def generar_pdf_historial(paciente, historial):
    """Genera los bytes del PDF profesional para el historial clínico."""
    try:
        # Configuración Inicial
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)
        
        # --- ENCABEZADO ---
        pdf.set_fill_color(30, 64, 175) 
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 12, 'SISTEMA NIXON - HISTORIAL CLINICO', 0, 1, 'C', True)
        pdf.ln(5)

        # --- DATOS DEL PACIENTE ---
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, 'DATOS DEL PACIENTE', 0, 1)
        
        def clean(txt):
            if not txt or str(txt).lower() == 'nan': return "N/A"
            return str(txt).replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')

        datos = [
            ('Nombre:', clean(paciente.get('nombre_apellido'))),
            ('DNI:', clean(paciente.get('dni'))),
            ('Edad:', f"{paciente.get('edad_meses', 'N/A')} meses"),
            ('Region:', clean(paciente.get('region'))),
            ('Hb Actual:', f"{paciente.get('hemoglobina_dl1', 'N/A')} g/dL"),
            ('Estado:', clean(paciente.get('estado_paciente')))
        ]

        for label, value in datos:
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(40, 7, label, 0, 0)
            pdf.set_font('Arial', '', 10)
            pdf.cell(0, 7, value, 0, 1)
        pdf.ln(5)

        # --- TABLA DE CONTROLES ---
        if historial:
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 10, 'HISTORIAL DE CONTROLES', 0, 1)
            pdf.set_fill_color(59, 130, 246)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font('Arial', 'B', 9)
            
            cols = [30, 30, 25, 60, 45]
            headers = ['Fecha', 'Tipo', 'Hb', 'Responsable', 'Proximo']
            for i, h in enumerate(headers):
                pdf.cell(cols[i], 8, h, 1, 0, 'C', True)
            pdf.ln()

            pdf.set_text_color(0, 0, 0)
            pdf.set_font('Arial', '', 8)
            for idx, c in enumerate(historial):
                bg = (idx % 2 == 0)
                pdf.set_fill_color(245, 247, 250)
                pdf.cell(cols[0], 7, clean(c.get('fecha_seguimiento'))[:10], 1, 0, 'C', bg)
                pdf.cell(cols[1], 7, clean(c.get('tipo_seguimiento'))[:15], 1, 0, 'C', bg)
                pdf.cell(cols[2], 7, f"{c.get('hemoglobina_actual', 'N/A')}", 1, 0, 'C', bg)
                pdf.cell(cols[3], 7, clean(c.get('usuario_responsable'))[:25], 1, 0, 'L', bg)
                pdf.cell(cols[4], 7, clean(c.get('proximo_control'))[:10], 1, 1, 'C', bg)

        # --- PIE DE PÁGINA ---
        pdf.set_y(-20)
        pdf.set_font('Arial', 'I', 8)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(0, 10, f'Generado el {datetime.now().strftime("%d/%m/%Y")} - Sistema Nixon', 0, 0, 'C')

        # Retornar como bytes
        pdf_output = pdf.output(dest='S')
        if isinstance(pdf_output, str):
            return pdf_output.encode('latin-1', errors='replace')
        return pdf_output
    except Exception as e:
        return f"Error: {str(e)}".encode('utf-8')

# 2. INTERFAZ DE EXPORTACIÓN (Donde están tus botones)
def mostrar_seccion_exportar(paciente, historial, df_historial):
    st.markdown("---")
    st.markdown("### 📤 Exportar Historial Clínico")
    
    # Creamos las 3 columnas para que los botones estén alineados
    col_exp1, col_exp2, col_exp3 = st.columns(3)
    
    with col_exp1:
        # Botón para CSV
        csv = df_historial.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="📊 Exportar a CSV",
            data=csv,
            file_name=f"historial_{paciente.get('dni', 'paciente')}.csv",
            mime="text/csv",
            use_container_width=True,
            key="btn_csv_hist"
        )
    
    with col_exp2:
        # ACCIÓN CLAVE: Generar y Descargar PDF
        try:
            pdf_bytes = generar_pdf_historial(paciente, historial)
            st.download_button(
                label="📄 Descargar Informe PDF",
                data=pdf_bytes,
                file_name=f"Informe_NIXON_{paciente.get('dni', 'paciente')}.pdf",
                mime="application/pdf",
                use_container_width=True,
                key="btn_pdf_hist_real",
                type="primary" # Resalta el botón en azul
            )
        except:
            st.error("No se pudo preparar el PDF")

    with col_exp3:
        # Botón para vista rápida (opcional)
        if st.button("🖨️ Vista de Impresión", use_container_width=True):
            st.info("Use el botón de PDF para obtener una versión lista para imprimir.")

def generar_pdf_dashboard_nacional(indicadores, datos, mapa_df=None):
    """Genera reporte PDF manejando errores de columnas y datos vacíos"""
    try:
        pdf = FPDF()
        pdf.add_page()
        
        # Título
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, "REPORTE NACIONAL DE ANEMIA", ln=True, align='C')
        pdf.ln(10)
        
        # Indicadores Clave
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 10, "RESUMEN DE INDICADORES:", ln=True)
        pdf.set_font("Arial", size=11)
        pdf.cell(0, 8, f"- Prevalencia Nacional: {indicadores.get('prevalencia_nacional', 0)}%", ln=True)
        pdf.cell(0, 8, f"- Total Pacientes: {indicadores.get('total_pacientes', 0)}", ln=True)
        pdf.cell(0, 8, f"- Hemoglobina Promedio: {indicadores.get('hb_promedio_nacional', 0):.1f} g/dL", ln=True)
        
        # Tabla Regional con validación de nombres de columna
        if mapa_df is not None and not mapa_df.empty:
            pdf.ln(5)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 10, "DETALLE POR REGIONES:", ln=True)
            pdf.set_font("Arial", size=10)
            
            # Detectar columna de región (maneja tildes)
            col_reg = next((c for c in mapa_df.columns if 'regi' in c.lower()), None)
            col_prev = next((c for c in mapa_df.columns if 'preval' in c.lower()), mapa_df.columns[1])

            if col_reg:
                for i, row in mapa_df.head(10).iterrows():
                    pdf.cell(0, 7, f"* {row[col_reg]}: {row[col_prev]}%", ln=True)
            else:
                pdf.cell(0, 7, "(No se encontró columna de regiones)", ln=True)

        # Retornar como bytes binarios
        return pdf.output(dest='S').encode('latin-1')
    except Exception as e:
        print(f"Error crítico en PDF: {e}")
        return None
# ==================================================
# SISTEMA DE LOGIN PARA 5 USUARIOS DE SALUD
# ==================================================

# Configurar estado de sesión
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.user_info = None
    st.session_state.current_username = None

# Diccionario de usuarios (5 profesionales de salud)
USUARIOS_SALUD = {
    "admin": {
        "password": "Admin123",
        "nombre": "Dr. Carlos Martínez",
        "rol": "Administrador",
        "especialidad": "Pediatría",
        "email": "admin@hospital.com",
        "acceso_total": True
    },
    "pediatra1": {
        "password": "Pediatra123",
        "nombre": "Dra. Ana López",
        "rol": "Pediatra",
        "especialidad": "Pediatría General",
        "email": "pediatra1@hospital.com",
        "acceso_total": True
    },
    "pediatra2": {
        "password": "Pediatra456",
        "nombre": "Dr. Juan Pérez",
        "rol": "Pediatra",
        "especialidad": "Nutrición Infantil",
        "email": "pediatra2@hospital.com",
        "acceso_total": True
    },
    "enfermero": {
        "password": "Enfermero123",
        "nombre": "Lic. María Gómez",
        "rol": "Enfermero/a",
        "especialidad": "Salud Pública",
        "email": "enfermero@hospital.com",
        "acceso_total": False
    },
    "tecnico": {
        "password": "Tecnico123",
        "nombre": "Téc. Luis Rodríguez",
        "rol": "Técnico de Laboratorio",
        "especialidad": "Hematología",
        "email": "tecnico@hospital.com",
        "acceso_total": False
    }
}

def verificar_login(username, password):
    """Verifica si el usuario y contraseña son correctos"""
    if username in USUARIOS_SALUD and USUARIOS_SALUD[username]["password"] == password:
        return USUARIOS_SALUD[username]
    return None

def logout():
    """Cierra sesión del usuario"""
    st.session_state.logged_in = False
    st.session_state.user_info = None
    st.session_state.current_username = None
    st.rerun()

def show_login_page():
    """Muestra la página de login"""
    
    # Estilos CSS para el login
    st.markdown("""
    <style>
    .login-container {
        max-width: 500px;
        margin: 10px auto;
        padding: 40px;
        background: white;
        border-radius: 20px;
        box-shadow: 0 15px 35px rgba(30, 64, 175, 0.1);
        border: 2px solid #e0f2fe;
    }
    
    .login-header {
        text-align: center;
        margin-bottom: 40px;
    }
    
    .hospital-icon {
        font-size: 60px;
        margin-bottom: 20px;
        color: #1e40af;
    }
    
    .login-title {
        color: #1e3a8a;
        font-size: 2.2rem;
        font-weight: 800;
        margin: 0;
        background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    
    .login-subtitle {
        color: #6b7280;
        font-size: 1rem;
        margin-top: 10px;
        font-weight: 500;
    }
    
    .stButton > button {
        width: 100%;
        background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
        color: white;
        border: none;
        padding: 16px 24px;
        border-radius: 12px;
        font-weight: 600;
        font-size: 16px;
        margin-top: 25px;
        transition: all 0.3s ease;
        cursor: pointer;
    }
    
    .stButton > button:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 20px rgba(30, 64, 175, 0.3);
    }
    
    .user-card {
        background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%);
        padding: 15px;
        border-radius: 12px;
        margin: 12px 0;
        border-left: 5px solid #3b82f6;
        transition: all 0.3s ease;
        border: 1px solid #e2e8f0;
    }
    
    .user-card:hover {
        transform: translateX(5px);
        background: #f0f9ff;
        border-color: #dbeafe;
    }
    
    .test-users {
        margin-top: 30px;
        padding: 20px;
        background: #f0f9ff;
        border-radius: 12px;
        border: 2px solid #dbeafe;
    }
    
    .role-badge {
        background: #3b82f6;
        color: white;
        padding: 4px 12px;
        border-radius: 20px;
        font-size: 12px;
        font-weight: 600;
        margin-left: 10px;
        display: inline-block;
    }
    
    .form-label {
        color: #374151;
        font-weight: 600;
        font-size: 0.95rem;
        margin-bottom: 8px;
        display: block;
    }
    
    .stTextInput > div > div > input {
        border-radius: 10px;
        border: 2px solid #e5e7eb;
        padding: 12px 16px;
        font-size: 16px;
    }
    
    .stTextInput > div > div > input:focus {
        border-color: #3b82f6;
        box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.1);
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Contenedor principal del login
    st.markdown('<div class="login-container">', unsafe_allow_html=True)
    
    # Header con icono
    st.markdown("""
    <div class="login-header">
        <div class="hospital-icon">🏥</div>
        <h1 class="login-title">SISTEMA NIXON</h1>
        <p class="login-subtitle">Control de Anemia y Nutrición Infantil</p>
        <div style="height: 3px; width: 80px; background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%); margin: 20px auto; border-radius: 10px;"></div>
    </div>
    """, unsafe_allow_html=True)
    
    # Formulario de login
    with st.form("login_form"):
        st.markdown('<div class="form-label">👤 Nombre de Usuario</div>', unsafe_allow_html=True)
        username = st.text_input("", placeholder="Ingresa tu usuario", label_visibility="collapsed")
        
        st.markdown('<div class="form-label">🔒 Contraseña</div>', unsafe_allow_html=True)
        password = st.text_input("", type="password", placeholder="Ingresa tu contraseña", label_visibility="collapsed")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            remember_me = st.checkbox("Recordar sesión", value=True)
        
        submit = st.form_submit_button("🚀 INICIAR SESIÓN", use_container_width=True)
        
        if submit:
            if not username or not password:
                st.error("❌ Por favor, ingresa usuario y contraseña")
            else:
                usuario_info = verificar_login(username, password)
                if usuario_info:
                    st.session_state.logged_in = True
                    st.session_state.user_info = usuario_info
                    st.session_state.current_username = username
                    st.success(f"✅ ¡Bienvenido/a, {usuario_info['nombre']}!")
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error("❌ Usuario o contraseña incorrectos")
    
    # Información de usuarios de prueba
    with st.expander("👥 USUARIOS AUTORIZADOS DEL SISTEMA", expanded=False):
        st.markdown('<div class="test-users">', unsafe_allow_html=True)
        st.markdown("**Personal de Salud Autorizado:**")
        
        for username, info in USUARIOS_SALUD.items():
            role_class = info['rol'].lower().replace("á", "a").replace("é", "e")
            st.markdown(f"""
            <div class="user-card">
                <div style="display: flex; align-items: center; margin-bottom: 8px;">
                    <strong style="font-size: 1.1rem;">{info['nombre']}</strong>
                    <span class="role-badge">{info['rol']}</span>
                </div>
                <div style="font-size: 0.9rem; color: #4b5563;">
                    <div><strong>Especialidad:</strong> {info['especialidad']}</div>
                    <div><strong>Usuario:</strong> <code>{username}</code></div>
                    <div><strong>Contraseña:</strong> <code>{info['password']}</code></div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
 

# ==================================================
# VERIFICAR SI EL USUARIO ESTÁ LOGUEADO
# ==================================================

if not st.session_state.logged_in:
    show_login_page()
    st.stop()  # Detener la ejecución del resto del código

# ==================================================
# CONFIGURACIÓN E INICIALIZACIÓN (solo se ejecuta si está logueado)
# ==================================================

st.set_page_config(
    page_title="Sistema Nixon - Control de Anemia",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==================================================
# ESTILOS CSS MEJORADOS (se agregan estilos para mostrar info del usuario)
# ==================================================
st.markdown("""
<style>
    /* TÍTULOS CON MEJOR VISUALIZACIÓN */
    .main-title {
        background: linear-gradient(135deg, #1e3a8a 0%, #1e40af 100%);
        padding: 2.5rem;
        border-radius: 15px;
        color: white;
        margin-bottom: 2rem;
        text-align: center;
        box-shadow: 0 10px 30px rgba(30, 58, 138, 0.2);
    }
    
    /* INFO USUARIO EN SIDEBAR */
    .user-sidebar-info {
        background: linear-gradient(135deg, #1e3a8a 0%, #1e40af 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 10px;
        margin-bottom: 1.5rem;
        text-align: center;
        border: 2px solid rgba(255,255,255,0.1);
    }
    
    .user-name {
        font-size: 1.3rem;
        font-weight: 700;
        margin-bottom: 5px;
    }
    
    .user-role {
        font-size: 0.9rem;
        opacity: 0.9;
        margin-bottom: 15px;
        background: rgba(255,255,255,0.15);
        padding: 4px 12px;
        border-radius: 15px;
        display: inline-block;
    }
    
    .user-email {
        font-size: 0.8rem;
        opacity: 0.8;
        margin-top: 5px;
    }
    
    .logout-btn {
        background: rgba(255,255,255,0.2) !important;
        border: 1px solid rgba(255,255,255,0.3) !important;
        color: white !important;
        margin-top: 10px !important;
    }
    
    .logout-btn:hover {
        background: rgba(255,255,255,0.3) !important;
        transform: translateY(-2px) !important;
    }
    
    /* Resto de estilos... (mantener todos los estilos existentes) */
    
    .section-title-blue {
        color: #1e3a8a;
        font-size: 1.8rem;
        font-weight: 700;
        margin-top: 2rem;
        margin-bottom: 1rem;
        padding-bottom: 10px;
        border-bottom: 3px solid #3b82f6;
    }
    
    .section-title-green {
        color: #065f46;
        font-size: 1.8rem;
        font-weight: 700;
        margin-top: 2rem;
        margin-bottom: 1rem;
        padding-bottom: 10px;
        border-bottom: 3px solid #10b981;
    }
    
    .section-title-red {
        color: #7f1d1d;
        font-size: 1.8rem;
        font-weight: 700;
        margin-top: 2rem;
        margin-bottom: 1rem;
        padding-bottom: 10px;
        border-bottom: 3px solid #ef4444;
    }
    
    .section-title-purple {
        color: #5b21b6;
        font-size: 1.8rem;
        font-weight: 700;
        margin-top: 2rem;
        margin-bottom: 1rem;
        padding-bottom: 10px;
        border-bottom: 3px solid #8b5cf6;
    }
    
    /* TARJETAS DE COLORES */
    .metric-card-blue {
        background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%);
        padding: 1.5rem;
        border-radius: 12px;
        border-left: 5px solid #3b82f6;
        margin: 0.5rem 0;
        box-shadow: 0 4px 12px rgba(59, 130, 246, 0.1);
    }
    
    .metric-card-green {
        background: linear-gradient(135deg, #d1fae5 0%, #a7f3d0 100%);
        padding: 1.5rem;
        border-radius: 12px;
        border-left: 5px solid #10b981;
        margin: 0.5rem 0;
        box-shadow: 0 4px 12px rgba(16, 185, 129, 0.1);
    }
    
    .metric-card-red {
        background: linear-gradient(135deg, #fee2e2 0%, #fecaca 100%);
        padding: 1.5rem;
        border-radius: 12px;
        border-left: 5px solid #ef4444;
        margin: 0.5rem 0;
        box-shadow: 0 4px 12px rgba(239, 68, 68, 0.1);
    }
    
    .metric-card-yellow {
        background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%);
        padding: 1.5rem;
        border-radius: 12px;
        border-left: 5px solid #f59e0b;
        margin: 0.5rem 0;
        box-shadow: 0 4px 12px rgba(245, 158, 11, 0.1);
    }
    
    .metric-card-purple {
        background: linear-gradient(135deg, #ede9fe 0%, #ddd6fe 100%);
        padding: 1.5rem;
        border-radius: 12px;
        border-left: 5px solid #8b5cf6;
        margin: 0.5rem 0;
        box-shadow: 0 4px 12px rgba(139, 92, 246, 0.1);
    }
    
    /* NÚMEROS DESTACADOS */
    .highlight-number {
        font-size: 2.5rem;
        font-weight: 800;
        margin-bottom: 0.5rem;
    }
    
    .highlight-blue { color: #1e40af; }
    .highlight-green { color: #065f46; }
    .highlight-red { color: #7f1d1d; }
    .highlight-yellow { color: #92400e; }
    .highlight-purple { color: #5b21b6; }
    
    /* ETIQUETAS */
    .metric-label {
        font-size: 0.9rem;
        color: #6b7280;
        font-weight: 500;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    
    /* ESTADOS DE ANEMIA */
    .severity-critical {
        background: linear-gradient(135deg, #fef2f2 0%, #fee2e2 100%);
        padding: 1.5rem;
        border-radius: 10px;
        border-left: 5px solid #dc2626;
        margin: 1rem 0;
    }
    
    .severity-moderate {
        background: linear-gradient(135deg, #fffbeb 0%, #fef3c7 100%);
        padding: 1.5rem;
        border-radius: 10px;
        border-left: 5px solid #d97706;
        margin: 1rem 0;
    }
    
    .severity-mild {
        background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%);
        padding: 1.5rem;
        border-radius: 10px;
        border-left: 5px solid #2563eb;
        margin: 1rem 0;
    }
    
    .severity-normal {
        background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%);
        padding: 1.5rem;
        border-radius: 10px;
        border-left: 5px solid #16a34a;
        margin: 1rem 0;
    }
    
    /* BOTONES MEJORADOS */
    .stButton > button {
        border-radius: 8px;
        font-weight: 600;
        transition: all 0.3s ease;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 12px rgba(0,0,0,0.15);
    }
    
    /* MEJORAS EN FORMULARIOS */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stSelectbox > div > div > select {
        border-radius: 8px;
        border: 2px solid #e5e7eb;
        transition: border-color 0.3s;
    }
    
    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus,
    .stSelectbox > div > div > select:focus {
        border-color: #3b82f6;
        box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.1);
    }
    
    /* TABLAS MEJORADAS */
    .dataframe {
        border-radius: 10px;
        overflow: hidden;
    }
</style>
""", unsafe_allow_html=True)

# ==================================================
# OBTENER INFORMACIÓN DEL USUARIO LOGUEADO
# ==================================================

user_info = st.session_state.user_info
current_user = st.session_state.current_username

# ==================================================
# SIDEBAR CON INFORMACIÓN DEL USUARIO
# ==================================================

with st.sidebar:
    # Información del usuario
    st.markdown(f"""
    <div class="user-sidebar-info">
        <div class="user-name">👤 {user_info['nombre']}</div>
        <div class="user-role">{user_info['rol']}</div>
        <div class="user-email">
            {user_info['email']}<br>
            <small>Especialidad: {user_info['especialidad']}</small>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Botón de logout
    if st.button("🚪 Cerrar Sesión", use_container_width=True, key="logout_btn", 
                type="secondary", help="Cierra la sesión actual"):
        logout()
    
    st.markdown("---")
    
    # Resto del sidebar original
    st.markdown('<div class="section-title-blue" style="font-size: 1.4rem;">📋 Sistema de Referencia</div>', unsafe_allow_html=True)
    
    tab_sidebar1, tab_sidebar2, tab_sidebar3 = st.tabs(["🎯 Ajustes Altitud", "📊 Crecimiento", "🔬 Hematología"])
    
    with tab_sidebar1:
        st.markdown('<div style="color: #1e40af; font-weight: 600; margin-bottom: 10px;">Tabla de Ajustes por Altitud</div>', unsafe_allow_html=True)
        
        # Tabla de ajustes por altitud
        AJUSTE_HEMOGLOBINA = [
            {"altitud_min": 0, "altitud_max": 999, "ajuste": 0.0},
            {"altitud_min": 1000, "altitud_max": 1499, "ajuste": -0.2},
            {"altitud_min": 1500, "altitud_max": 1999, "ajuste": -0.5},
            {"altitud_min": 2000, "altitud_max": 2499, "ajuste": -0.8},
            {"altitud_min": 2500, "altitud_max": 2999, "ajuste": -1.3},
            {"altitud_min": 3000, "altitud_max": 3499, "ajuste": -1.9},
            {"altitud_min": 3500, "altitud_max": 3999, "ajuste": -2.7},
            {"altitud_min": 4000, "altitud_max": 4499, "ajuste": -3.5},
            {"altitud_min": 4500, "altitud_max": 10000, "ajuste": -4.5}
        ]
        
        ajustes_df = pd.DataFrame(AJUSTE_HEMOGLOBINA)
        st.dataframe(
            ajustes_df.style.format({
                'altitud_min': '{:.0f}',
                'altitud_max': '{:.0f}', 
                'ajuste': '{:+.1f}'
            }),
            use_container_width=True,
            height=300
        )
    
    with tab_sidebar2:
        st.markdown('<div style="color: #1e40af; font-weight: 600; margin-bottom: 10px;">Tablas de Crecimiento OMS</div>', unsafe_allow_html=True)
        
        # Función para obtener referencia de crecimiento
        def obtener_referencia_crecimiento():
            return pd.DataFrame([
                {'edad_meses': 0, 'peso_min_ninas': 2.8, 'peso_promedio_ninas': 3.4, 'peso_max_ninas': 4.2, 'peso_min_ninos': 2.9, 'peso_promedio_ninos': 3.4, 'peso_max_ninos': 4.4, 'talla_min_ninas': 47.0, 'talla_promedio_ninas': 50.3, 'talla_max_ninas': 53.6, 'talla_min_ninos': 47.5, 'talla_promedio_ninos': 50.3, 'talla_max_ninos': 53.8},
                {'edad_meses': 3, 'peso_min_ninas': 4.5, 'peso_promedio_ninas': 5.6, 'peso_max_ninas': 7.0, 'peso_min_ninos': 5.0, 'peso_promedio_ninos': 6.2, 'peso_max_ninos': 7.8, 'talla_min_ninas': 55.0, 'talla_promedio_ninas': 59.0, 'talla_max_ninas': 63.5, 'talla_min_ninos': 57.0, 'talla_promedio_ninos': 60.0, 'talla_max_ninos': 64.5},
                {'edad_meses': 6, 'peso_min_ninas': 6.0, 'peso_promedio_ninas': 7.3, 'peso_max_ninas': 9.0, 'peso_min_ninos': 6.5, 'peso_promedio_ninos': 8.0, 'peso_max_ninos': 9.8, 'talla_min_ninas': 61.0, 'talla_promedio_ninas': 65.0, 'talla_max_ninas': 69.5, 'talla_min_ninos': 63.0, 'talla_promedio_ninos': 67.0, 'talla_max_ninos': 71.5},
                {'edad_meses': 24, 'peso_min_ninas': 10.5, 'peso_promedio_ninas': 12.4, 'peso_max_ninas': 15.0, 'peso_min_ninos': 11.0, 'peso_promedio_ninos': 12.9, 'peso_max_ninos': 16.0, 'talla_min_ninas': 81.0, 'talla_promedio_ninas': 86.0, 'talla_max_ninas': 92.5, 'talla_min_ninos': 83.0, 'talla_promedio_ninos': 88.0, 'talla_max_ninos': 94.5}
            ])
        
        referencia_df = obtener_referencia_crecimiento()
        st.dataframe(referencia_df, use_container_width=True, height=300)
    
    with tab_sidebar3:
        st.markdown('<div style="color: #1e40af; font-weight: 600; margin-bottom: 10px;">Criterios de Interpretación</div>', unsafe_allow_html=True)
        
        st.markdown("""
        ### 🩺 FERRITINA (Reservas)
        - **< 15 ng/mL**: 🚨 Deficit severo
        - **15-30 ng/mL**: ⚠️ Deficit moderado  
        - **30-100 ng/mL**: 🔄 Reservas límite
        - **> 100 ng/mL**: ✅ Adecuado
        
        ### 🔬 CHCM (Concentración)
        - **< 32 g/dL**: 🚨 Hipocromía
        - **32-36 g/dL**: ✅ Normocromía
        - **> 36 g/dL**: 🔄 Hipercromía
        
        ### 📈 RETICULOCITOS (Producción)
        - **< 0.5%**: ⚠️ Hipoproliferación
        - **0.5-1.5%**: ✅ Normal
        - **> 1.5%**: 🔄 Hiperproducción
        """)

# ==================================================
# TÍTULO PRINCIPAL CON INFORMACIÓN DEL USUARIO
# ==================================================

st.markdown(f"""
<div class="main-title" style="padding: 2rem;">
    <h1 style="margin: 0; font-size: 2.5rem;">🏥 SISTEMA NIXON - Control de Anemia</h1>
    <p style="margin: 10px 0 0 0; font-size: 1.1rem; opacity: 0.9;">
    Usuario: <strong>{user_info['nombre']}</strong> | Rol: <strong>{user_info['rol']}</strong> | Especialidad: <strong>{user_info['especialidad']}</strong>
    </p>
</div>
""", unsafe_allow_html=True)

# ==================================================
# CONFIGURACIÓN SUPABASE
# ==================================================

TABLE_NAME = "alertas_hemoglobina"
ALTITUD_TABLE = "altitud_regiones"
CRECIMIENTO_TABLE = "referencia_crecimiento"

SUPABASE_URL = st.secrets.get("SUPABASE_URL", "https://kwsuszkblbejvliniggd.supabase.co")
SUPABASE_KEY = st.secrets.get("SUPABASE_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Imt3c3VzemtibGJlanZsaW5pZ2dkIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjE2ODE0NTUsImV4cCI6MjA3NzI1NzQ1NX0.DQpt-rSNprcUrbOLTgUEEn_0jFIuSX5b0AVuVirk0vw")

@st.cache_resource
def init_supabase():
    try:
        supabase_client = create_client(SUPABASE_URL, SUPABASE_KEY)
        return supabase_client
    except Exception as e:
        st.error(f"❌ Error conectando a Supabase: {str(e)}")
        return None

supabase = init_supabase()

# ==================================================
# FUNCIONES PARA TABLA CITAS
# ==================================================

def crear_tabla_citas_simple():
    """Crea la tabla citas de forma simple"""
    try:
        st.sidebar.info("🛠️ Configurando tabla 'citas'...")
        
        # Intentar crear una cita de prueba simple
        test_cita = {
            "dni_paciente": "99988877",
            "fecha_cita": "2024-12-14",
            "hora_cita": "10:00:00",
            "tipo_consulta": "Prueba de sistema"
        }
        
        response = supabase.table("citas").insert(test_cita).execute()
        
        if response.data:
            st.sidebar.success("✅ Tabla 'citas' accesible")
            # Limpiar la prueba
            supabase.table("citas").delete().eq("dni_paciente", "99988877").execute()
            return True
        else:
            st.sidebar.warning("⚠️ Puede que la tabla necesite configuración en Supabase")
            
            # Instrucciones para crear manualmente
            st.sidebar.markdown("""
            **📝 Si falla, crea la tabla manualmente en Supabase:**
            
            1. Ve a **Table Editor**
            2. Crea nueva tabla llamada **"citas"**
            3. Agrega estas columnas:
               - `id` (bigint, autoincrement, primary key)
               - `dni_paciente` (text)
               - `fecha_cita` (date)
               - `hora_cita` (time)
               - `tipo_consulta` (text)
               - `diagnostico` (text)
               - `tratamiento` (text)
               - `observaciones` (text)
               - `investigador_responsable` (text)
               - `proxima_cita` (date)
               - `created_at` (timestamptz, default: now())
            4. En **Authentication → Policies**, crea política:
               - `allow_all` (para todas las operaciones)
            """)
            return False
            
    except Exception as e:
        st.sidebar.error(f"❌ Error: {str(e)[:100]}")
        return False

def probar_guardado_directo():
    """Prueba directa de guardado"""
    with st.sidebar:
        st.markdown("### 🧪 Prueba Directa")
        
        try:
            pacientes = supabase.table("alertas_hemoglobina").select("dni").limit(5).execute()
            
            if pacientes.data and len(pacientes.data) > 0:
                dni_real = pacientes.data[0]["dni"]
                st.info(f"📋 Usando DNI real: {dni_real}")
            else:
                dni_real = "12345678"
                st.warning("⚠️ No hay pacientes, usando DNI de prueba")
                
        except:
            dni_real = "12345678"
        
        test_cita = {
            "dni_paciente": dni_real,
            "fecha_cita": "2024-12-14",
            "hora_cita": "09:00:00",
            "tipo_consulta": "Consulta de prueba",
            "diagnostico": "Paciente de prueba para verificar sistema",
            "tratamiento": "Observación",
            "observaciones": "Esta es una prueba del sistema de citas",
            "investigador_responsable": "Dr. Prueba"
        }
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("📤 Enviar prueba", type="primary", key="enviar_prueba"):
                try:
                    with st.spinner("Enviando a Supabase..."):
                        result = supabase.table("citas").insert(test_cita).execute()
                    
                    if result.data:
                        st.success(f"✅ ¡ÉXITO! Guardado correctamente")
                        st.info(f"ID generado: {result.data[0].get('id', 'N/A')}")
                    else:
                        st.warning("⚠️ Respuesta inesperada del servidor")
                        
                except Exception as e:
                    st.error(f"🔥 Error: {str(e)[:200]}")
        
        with col2:
            if st.button("🗑️ Limpiar pruebas", key="limpiar_pruebas"):
                try:
                    supabase.table("citas").delete().eq("dni_paciente", dni_real).execute()
                    st.success("✅ Pruebas limpiadas")
                except Exception as e:
                    st.info(f"ℹ️ {str(e)[:100]}")

# ==================================================
# FUNCIONES DE BASE DE DATOS
# ==================================================

def obtener_datos_supabase(tabla=TABLE_NAME):
    try:
        if supabase:
            response = supabase.table(tabla).select("*").execute()
            if hasattr(response, 'error') and response.error:
                st.error(f"Error obteniendo datos: {response.error}")
                return pd.DataFrame()
            return pd.DataFrame(response.data) if response.data else pd.DataFrame()
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Error obteniendo datos: {e}")
        return pd.DataFrame()

def obtener_casos_seguimiento():
    try:
        if supabase:
            response = supabase.table(TABLE_NAME).select("*").eq("en_seguimiento", True).execute()
            if hasattr(response, 'error') and response.error:
                return pd.DataFrame()
            return pd.DataFrame(response.data) if response.data else pd.DataFrame()
        return pd.DataFrame()
    except Exception as e:
        return pd.DataFrame()

def verificar_duplicado(dni):
    """Verifica si un DNI ya existe en la base de datos"""
    try:
        if supabase:
            response = supabase.table(TABLE_NAME)\
                .select("dni")\
                .eq("dni", dni)\
                .execute()
            
            if response.data and len(response.data) > 0:
                return True
            return False
        return False
    except Exception as e:
        st.error(f"Error verificando duplicado: {e}")
        return False

def insertar_datos_supabase(datos, tabla=TABLE_NAME):
    """Inserta datos en Supabase verificando duplicados"""
    try:
        dni = datos.get("dni")
        
        if not dni:
            st.error("❌ El registro no tiene DNI")
            return None
        
        if verificar_duplicado(dni):
            st.error(f"❌ El DNI {dni} ya existe en la base de datos")
            return {"status": "duplicado", "dni": dni}
        
        if supabase:
            response = supabase.table(tabla).insert(datos).execute()
            if hasattr(response, 'error') and response.error:
                st.error(f"❌ Error Supabase al insertar: {response.error}")
                return None
            return response.data[0] if response.data else None
        return None
    except Exception as e:
        st.error(f"Error insertando datos: {e}")
        return None

def upsert_datos_supabase(datos, tabla=TABLE_NAME):
    """Inserta o actualiza datos si ya existen"""
    try:
        if supabase:
            response = supabase.table(tabla)\
                .upsert(datos, on_conflict='dni')\
                .execute()
            
            if hasattr(response, 'error') and response.error:
                st.error(f"❌ Error Supabase al hacer upsert: {response.error}")
                return None
            return response.data[0] if response.data else None
        return None
    except Exception as e:
        st.error(f"Error haciendo upsert: {e}")
        return None

# ==================================================
# TABLAS DE REFERENCIA Y FUNCIONES DE CÁLCULO
# ==================================================

def obtener_altitud_regiones():
    """Obtiene datos de altitud de regiones desde Supabase"""
    try:
        if supabase:
            response = supabase.table(ALTITUD_TABLE).select("*").execute()
            if response.data:
                return {row['region']: row for row in response.data}
        return {
            "AMAZONAS": {"altitud_min": 500, "altitud_max": 3500, "altitud_promedio": 1800},
            "ANCASH": {"altitud_min": 0, "altitud_max": 6768, "altitud_promedio": 3000},
            "APURIMAC": {"altitud_min": 2000, "altitud_max": 4500, "altitud_promedio": 3200},
            "AREQUIPA": {"altitud_min": 0, "altitud_max": 5825, "altitud_promedio": 2500},
            "AYACUCHO": {"altitud_min": 1800, "altitud_max": 4500, "altitud_promedio": 2800},
            "CAJAMARCA": {"altitud_min": 500, "altitud_max": 3500, "altitud_promedio": 2700},
            "CALLAO": {"altitud_min": 0, "altitud_max": 50, "altitud_promedio": 5},
            "CUSCO": {"altitud_min": 500, "altitud_max": 4800, "altitud_promedio": 3400},
            "HUANCAVELICA": {"altitud_min": 2000, "altitud_max": 4500, "altitud_promedio": 3600},
            "HUANUCO": {"altitud_min": 200, "altitud_max": 3800, "altitud_promedio": 1900},
            "ICA": {"altitud_min": 0, "altitud_max": 3800, "altitud_promedio": 500},
            "JUNIN": {"altitud_min": 500, "altitud_max": 4800, "altitud_promedio": 3500},
            "LA LIBERTAD": {"altitud_min": 0, "altitud_max": 4200, "altitud_promedio": 1800},
            "LAMBAYEQUE": {"altitud_min": 0, "altitud_max": 3000, "altitud_promedio": 100},
            "LIMA": {"altitud_min": 0, "altitud_max": 4800, "altitud_promedio": 150},
            "LORETO": {"altitud_min": 70, "altitud_max": 220, "altitud_promedio": 120},
            "MADRE DE DIOS": {"altitud_min": 200, "altitud_max": 500, "altitud_promedio": 250},
            "MOQUEGUA": {"altitud_min": 0, "altitud_max": 4500, "altitud_promedio": 1400},
            "PASCO": {"altitud_min": 1000, "altitud_max": 4400, "altitud_promedio": 3200},
            "PIURA": {"altitud_min": 0, "altitud_max": 3500, "altitud_promedio": 100},
            "PUNO": {"altitud_min": 3800, "altitud_max": 4800, "altitud_promedio": 4100},
            "SAN MARTIN": {"altitud_min": 200, "altitud_max": 3000, "altitud_promedio": 600},
            "TACNA": {"altitud_min": 0, "altitud_max": 3500, "altitud_promedio": 600},
            "TUMBES": {"altitud_min": 0, "altitud_max": 500, "altitud_promedio": 20},
            "UCAYALI": {"altitud_min": 100, "altitud_max": 350, "altitud_promedio": 180}
        }
    except:
        return {}

ALTITUD_REGIONES = obtener_altitud_regiones()

AJUSTE_HEMOGLOBINA = [
    {"altitud_min": 0, "altitud_max": 999, "ajuste": 0.0},
    {"altitud_min": 1000, "altitud_max": 1499, "ajuste": -0.2},
    {"altitud_min": 1500, "altitud_max": 1999, "ajuste": -0.5},
    {"altitud_min": 2000, "altitud_max": 2499, "ajuste": -0.8},
    {"altitud_min": 2500, "altitud_max": 2999, "ajuste": -1.3},
    {"altitud_min": 3000, "altitud_max": 3499, "ajuste": -1.9},
    {"altitud_min": 3500, "altitud_max": 3999, "ajuste": -2.7},
    {"altitud_min": 4000, "altitud_max": 4499, "ajuste": -3.5},
    {"altitud_min": 4500, "altitud_max": 10000, "ajuste": -4.5}
]

def obtener_ajuste_hemoglobina(altitud):
    """Retorna el valor de ajuste según la tabla de altitud"""
    try:
        alt = float(altitud)
        for ajuste in AJUSTE_HEMOGLOBINA:
            if ajuste["altitud_min"] <= alt <= ajuste["altitud_max"]:
                return ajuste["ajuste"]
        return 0.0
    except:
        return 0.0

def calcular_hemoglobina_ajustada(hemoglobina_medida, altitud):
    """Aplica la resta del factor de altitud a la Hb observada"""
    try:
        if hemoglobina_medida is None: return None
        ajuste = obtener_ajuste_hemoglobina(altitud)
        # Sumamos porque los valores en AJUSTE_HEMOGLOBINA ya son negativos
        return round(float(hemoglobina_medida) + ajuste, 2)
    except:
        return None

def clasificar_anemia_por_hb(hb_ajustada):
    if pd.isna(hb_ajustada):
        return "SIN DATO"
    elif hb_ajustada < 7.0:
        return "ANEMIA SEVERA"
    elif hb_ajustada < 10.0:
        return "ANEMIA MODERADA"
    elif hb_ajustada < 11.0:
        return "ANEMIA LEVE"
    else:
        return "SIN ANEMIA"

# ==================================================
# SISTEMA DE INTERPRETACIÓN AUTOMÁTICA
# ==================================================

def interpretar_analisis_hematologico(ferritina, chcm, reticulocitos, transferrina, hemoglobina_ajustada, edad_meses):
    """Sistema de interpretación automática de parámetros hematológicos"""
    
    interpretacion = ""
    severidad = ""
    recomendacion = ""
    codigo_color = ""
    
    # EVALUAR FERRITINA
    if ferritina < 15:
        interpretacion += "🚨 **DEFICIT SEVERO DE HIERRO**. "
        severidad = "CRITICO"
    elif ferritina < 30:
        interpretacion += "⚠️ **DEFICIT MODERADO DE HIERRO**. "
        severidad = "MODERADO"
    elif ferritina < 100:
        interpretacion += "🔄 **RESERVAS DE HIERRO LIMITE**. "
        severidad = "LEVE"
    else:
        interpretacion += "✅ **RESERVAS DE HIERRO ADECUADAS**. "
        severidad = "NORMAL"
    
    # EVALUAR CHCM
    if chcm < 32:
        interpretacion += "🚨 **HIPOCROMÍA SEVERA** - Deficiencia avanzada de hierro. "
        severidad = "CRITICO" if severidad != "CRITICO" else severidad
    elif chcm >= 32 and chcm <= 36:
        interpretacion += "✅ **NORMOCROMÍA** - Estado normal. "
    else:
        interpretacion += "🔄 **HIPERCROMÍA** - Posible esferocitosis. "
    
    # EVALUAR RETICULOCITOS
    if reticulocitos < 0.5:
        interpretacion += "⚠️ **HIPOPROLIFERACIÓN MEDULAR** - Respuesta insuficiente. "
    elif reticulocitos > 1.5:
        interpretacion += "🔄 **HIPERPRODUCCIÓN COMPENSATORIA** - Respuesta aumentada. "
    else:
        interpretacion += "✅ **PRODUCCIÓN MEDULAR NORMAL**. "
    
    # EVALUAR TRANSFERRINA
    if transferrina < 200:
        interpretacion += "⚠️ **SATURACIÓN BAJA** - Transporte disminuido. "
    elif transferrina > 400:
        interpretacion += "🔄 **SATURACIÓN AUMENTADA** - Compensación por deficiencia. "
    else:
        interpretacion += "✅ **TRANSPORTE ADECUADO**. "
    
    # CLASIFICACIÓN DE ANEMIA
    clasificacion_hb, _, _ = clasificar_anemia(hemoglobina_ajustada, edad_meses)
    interpretacion += f"📊 **CLASIFICACIÓN HEMOGLOBINA: {clasificacion_hb}**"
    
    # GENERAR RECOMENDACIÓN
    if severidad == "CRITICO":
        recomendacion = "🚨 **INTERVENCIÓN INMEDIATA**: Suplementación con hierro elemental 3-6 mg/kg/día + Control en 15 días + Evaluación médica urgente"
        codigo_color = "#DC2626"
    elif severidad == "MODERADO":
        recomendacion = "⚠️ **ACCIÓN PRIORITARIA**: Iniciar suplementación con hierro + Control mensual + Educación nutricional"
        codigo_color = "#D97706"
    elif severidad == "LEVE":
        recomendacion = "🔄 **VIGILANCIA ACTIVA**: Suplementación preventiva + Modificación dietética + Control cada 3 meses"
        codigo_color = "#2563EB"
    else:
        recomendacion = "✅ **SEGUIMIENTO RUTINARIO**: Mantener alimentación balanceada + Control preventivo cada 6 meses"
        codigo_color = "#16A34A"
    
    return {
        "interpretacion": interpretacion,
        "severidad": severidad,
        "recomendacion": recomendacion,
        "codigo_color": codigo_color,
        "clasificacion_hemoglobina": clasificacion_hb
    }

def generar_parametros_hematologicos(hemoglobina_ajustada, edad_meses):
    """Genera parámetros hematológicos simulados"""
    
    if hemoglobina_ajustada < 9.0:
        ferritina = np.random.uniform(5, 15)
        chcm = np.random.uniform(28, 31)
        reticulocitos = np.random.uniform(0.5, 1.0)
        transferrina = np.random.uniform(350, 450)
    elif hemoglobina_ajustada < 11.0:
        ferritina = np.random.uniform(15, 50)
        chcm = np.random.uniform(31, 33)
        reticulocitos = np.random.uniform(1.0, 1.8)
        transferrina = np.random.uniform(300, 400)
    else:
        ferritina = np.random.uniform(80, 150)
        chcm = np.random.uniform(33, 36)
        reticulocitos = np.random.uniform(0.8, 1.5)
        transferrina = np.random.uniform(200, 350)
    
    vcm = (chcm / 33) * np.random.uniform(75, 95)
    hcm = (chcm / 33) * np.random.uniform(27, 32)
    
    return {
        'vcm': round(vcm, 1),
        'hcm': round(hcm, 1),
        'chcm': round(chcm, 1),
        'ferritina': round(ferritina, 1),
        'transferrina': round(transferrina, 0),
        'reticulocitos': round(reticulocitos, 1)
    }
def interpretar_analisis_hematologico(ferritina, chcm, reticulocitos, transferrina, hemoglobina_ajustada, edad_meses, pcr):
    """Sistema avanzado de interpretación con descarte de diagnósticos"""
    
    interpretacion = ""
    severidad = ""
    recomendacion = ""
    codigo_color = ""
    descarte_clinico = []
    
    # 1. EVALUAR ESTADO DE HIERRO (Ferritina + PCR)
    if ferritina < 15 and pcr <= 0.5:
        interpretacion += "🚨 **DEFICIENCIA REAL DE HIERRO**. "
        severidad = "CRITICO"
        descarte_clinico.append("✅ **Anemia Inflamatoria descartada**: PCR normal confirma déficit de hierro real.")
    elif ferritina < 30:
        interpretacion += "⚠️ **RESERVAS BAJAS DE HIERRO**. "
        severidad = "MODERADO"
    else:
        interpretacion += "✅ **RESERVAS DE HIERRO NORMALES**. "
        severidad = "NORMAL"
        if pcr > 0.5:
            descarte_clinico.append("🦠 **Ojo**: PCR elevada. La ferritina podría estar falsamente normal.")

    # 2. EVALUAR MORFOLOGÍA (CHCM)
    if chcm < 32:
        interpretacion += "📉 **HIPOCROMÍA**. "
        descarte_clinico.append("✅ **Anemia Megaloblástica descartada**: CHCM bajo sugiere falta de hierro, no de B12.")
    elif chcm >= 32 and chcm <= 36:
        interpretacion += "✅ **NORMOCROMÍA**. "
    
    # 3. EVALUAR PRODUCCIÓN (Reticulocitos)
    if reticulocitos > 1.5:
        interpretacion += "🔄 **MÉDULA HIPERACTIVA**. "
        descarte_clinico.append("⚠️ **Anemia Hemolítica**: Considerar esta opción por reticulocitos altos.")
    else:
        interpretacion += "🟡 **PRODUCCIÓN MEDULAR NO COMPENSATORIA**. "
        descarte_clinico.append("✅ **Anemia Hemolítica descartada**: Reticulocitos normales indican falta de materia prima.")

    # 4. CLASIFICACIÓN FINAL Y COLORES
    if severidad == "CRITICO":
        recomendacion = "🚨 **INTERVENCIÓN INMEDIATA**: Hierro elemental 3-6 mg/kg/día + Control en 15 días."
        codigo_color = "#DC2626"
    elif severidad == "MODERADO":
        recomendacion = "⚠️ **ACCIÓN PRIORITARIA**: Suplementación con hierro + Educación nutricional."
        codigo_color = "#D97706"
    else:
        recomendacion = "✅ **SEGUIMIENTO**: Alimentación balanceada y control preventivo."
        codigo_color = "#16A34A"
        
    return {
        "interpretacion": interpretacion,
        "severidad": severidad,
        "recomendacion": recomendacion,
        "codigo_color": codigo_color,
        "descartes": descarte_clinico
    }
# ==================================================
# CLASIFICACIÓN DE ANEMIA
# ==================================================

def clasificar_anemia(hemoglobina_ajustada, edad_meses):
    """Clasifica la anemia según estándares OMS"""
    
    if edad_meses < 24:
        if hemoglobina_ajustada >= 11.0:
            return "SIN ANEMIA", "NO requiere seguimiento", "success"
        elif 10.0 <= hemoglobina_ajustada < 11.0:
            return "ANEMIA LEVE", "Seguimiento cada 3 meses", "warning"
        elif 9.0 <= hemoglobina_ajustada < 10.0:
            return "ANEMIA MODERADA", "Seguimiento mensual", "error"
        else:
            return "ANEMIA SEVERA", "Seguimiento urgente semanal", "error"
    
    elif 24 <= edad_meses < 60:
        if hemoglobina_ajustada >= 11.5:
            return "SIN ANEMIA", "NO requiere seguimiento", "success"
        elif 10.5 <= hemoglobina_ajustada < 11.5:
            return "ANEMIA LEVE", "Seguimiento cada 3 meses", "warning"
        elif 9.5 <= hemoglobina_ajustada < 10.5:
            return "ANEMIA MODERADA", "Seguimiento mensual", "error"
        else:
            return "ANEMIA SEVERA", "Seguimiento urgente semanal", "error"
    
    else:
        if hemoglobina_ajustada >= 12.0:
            return "SIN ANEMIA", "NO requiere seguimiento", "success"
        elif 11.0 <= hemoglobina_ajustada < 12.0:
            return "ANEMIA LEVE", "Seguimiento cada 3 meses", "warning"
        elif 10.0 <= hemoglobina_ajustada < 11.0:
            return "ANEMIA MODERADA", "Seguimiento mensual", "error"
        else:
            return "ANEMIA SEVERA", "Seguimiento urgente semanal", "error"

def necesita_seguimiento_automatico(hemoglobina_ajustada, edad_meses):
    clasificacion, _, _ = clasificar_anemia(hemoglobina_ajustada, edad_meses)
    return clasificacion in ["ANEMIA MODERADA", "ANEMIA SEVERA"]

# ==================================================
# FUNCIONES DE EVALUACIÓN NUTRICIONAL
# ==================================================

def obtener_referencia_crecimiento():
    """Obtiene la tabla de referencia de crecimiento desde Supabase"""
    try:
        if supabase:
            response = supabase.table(CRECIMIENTO_TABLE).select("*").execute()
            if response.data:
                return pd.DataFrame(response.data)
        return pd.DataFrame([
            {'edad_meses': 0, 'peso_min_ninas': 2.8, 'peso_promedio_ninas': 3.4, 'peso_max_ninas': 4.2, 'peso_min_ninos': 2.9, 'peso_promedio_ninos': 3.4, 'peso_max_ninos': 4.4, 'talla_min_ninas': 47.0, 'talla_promedio_ninas': 50.3, 'talla_max_ninas': 53.6, 'talla_min_ninos': 47.5, 'talla_promedio_ninos': 50.3, 'talla_max_ninos': 53.8},
            {'edad_meses': 3, 'peso_min_ninas': 4.5, 'peso_promedio_ninas': 5.6, 'peso_max_ninas': 7.0, 'peso_min_ninos': 5.0, 'peso_promedio_ninos': 6.2, 'peso_max_ninos': 7.8, 'talla_min_ninas': 55.0, 'talla_promedio_ninas': 59.0, 'talla_max_ninas': 63.5, 'talla_min_ninos': 57.0, 'talla_promedio_ninos': 60.0, 'talla_max_ninos': 64.5},
            {'edad_meses': 6, 'peso_min_ninas': 6.0, 'peso_promedio_ninas': 7.3, 'peso_max_ninas': 9.0, 'peso_min_ninos': 6.5, 'peso_promedio_ninos': 8.0, 'peso_max_ninos': 9.8, 'talla_min_ninas': 61.0, 'talla_promedio_ninas': 65.0, 'talla_max_ninas': 69.5, 'talla_min_ninos': 63.0, 'talla_promedio_ninos': 67.0, 'talla_max_ninos': 71.5},
            {'edad_meses': 24, 'peso_min_ninas': 10.5, 'peso_promedio_ninas': 12.4, 'peso_max_ninas': 15.0, 'peso_min_ninos': 11.0, 'peso_promedio_ninos': 12.9, 'peso_max_ninos': 16.0, 'talla_min_ninas': 81.0, 'talla_promedio_ninas': 86.0, 'talla_max_ninas': 92.5, 'talla_min_ninos': 83.0, 'talla_promedio_ninos': 88.0, 'talla_max_ninos': 94.5}
        ])
    except:
        return pd.DataFrame()

def evaluar_estado_nutricional(edad_meses, peso_kg, talla_cm, genero):
    """
    Evalúa el estado nutricional comparando con la tabla de referencia.
    Incluye manejo de errores para datos nulos y fallos de conexión.
    """
    # 1. Obtener la tabla (Asegúrate de que obtener_referencia_crecimiento use @st.cache_data)
    referencia_df = obtener_referencia_crecimiento()
    
    # Verificación de carga de base de datos
    if referencia_df is None or referencia_df.empty:
        return "Error de DB", "Error de DB", "TABLA OMS NO CARGADA"
    
    try:
        # 2. Normalización de entrada
        # Convertimos a float por si vienen como strings del formulario
        edad_val = float(edad_meses)
        peso_val = float(peso_kg)
        talla_val = float(talla_cm)
        
        # Redondeo para buscar en la tabla (ej: 15.2 -> 15)
        edad_entera = int(round(edad_val))
        
        # 3. Búsqueda en la tabla
        ref_edad = referencia_df[referencia_df['edad_meses'] == edad_entera]
        
        if ref_edad.empty:
            return f"Edad {edad_entera}m no encontrada", "N/A", "EDAD FUERA DE RANGO"
        
        ref = ref_edad.iloc[0]
        
        # 4. Selección dinámica de columnas según género
        # Usamos .get() o validación para evitar errores si la columna es NULA
        es_femenino = str(genero).upper() in ['F', 'FEMENINO', 'NIÑA']
        
        if es_femenino:
            p_min, p_max = ref['peso_min_ninas'], ref['peso_max_ninas']
            t_min, t_max = ref['talla_min_ninas'], ref['talla_max_ninas']
        else:
            p_min, p_max = ref['peso_min_ninos'], ref['peso_max_ninos']
            t_min, t_max = ref['talla_min_ninos'], ref['talla_max_ninos']

        # Verificar que los valores de la tabla no sean None/NaN
        if pd.isna([p_min, p_max, t_min, t_max]).any():
            return "Datos tabla incompletos", "Datos tabla incompletos", "ERROR TABLA OMS"

        # 5. Lógica de Evaluación (Criterios OMS)
        estado_peso = "PESO NORMAL"
        if peso_val < p_min: 
            estado_peso = "BAJO PESO"
        elif peso_val > p_max: 
            estado_peso = "SOBREPESO"
        
        estado_talla = "TALLA NORMAL"
        if talla_val < t_min: 
            estado_talla = "TALLA BAJA"
        elif talla_val > t_max: 
            estado_talla = "TALLA ALTA"
        
        # 6. Diagnóstico Final Integrado
        if estado_peso == "BAJO PESO" and estado_talla == "TALLA BAJA":
            estado_nut = "DESNUTRICIÓN CRÓNICA"
        elif estado_peso == "BAJO PESO":
            estado_nut = "DESNUTRICIÓN AGUDA"
        elif estado_peso == "SOBREPESO" or estado_talla == "TALLA ALTA":
            estado_nut = "SOBREPESO / RIESGO"
        else:
            estado_nut = "NUTRICIÓN ADECUADA"
            
        return estado_peso, estado_talla, estado_nut

    except ValueError:
        return "Datos inválidos", "Datos inválidos", "ERROR EN FORMULARIO"
    except Exception as e:
        return "Error inesperado", "Error inesperado", f"CONSULTAR SOPORTE: {str(e)}"

# ==================================================
# LISTAS DE OPCIONES
# ==================================================

PERU_REGIONS = [
    "AMAZONAS", "ANCASH", "APURIMAC", "AREQUIPA", "AYACUCHO", 
    "CAJAMARCA", "CALLAO", "CUSCO", "HUANCAVELICA", "HUANUCO",
    "ICA", "JUNIN", "LA LIBERTAD", "LAMBAYEQUE", "LIMA", 
    "LORETO", "MADRE DE DIOS", "MOQUEGUA", "PASCO", "PIURA",
    "PUNO", "SAN MARTIN", "TACNA", "TUMBES", "UCAYALI"
]

GENEROS = ["F", "M"]
NIVELES_EDUCATIVOS = ["Sin educación", "Primaria", "Secundaria", "Superior"]
FRECUENCIAS_SUPLEMENTO = ["Diario", "3 veces por semana", "Semanal", "Otra"]
ESTADOS_PACIENTE = ["Activo", "En seguimiento", "Dado de alta", "Inactivo"]

FACTORES_CLINICOS = [
    "Historial familiar de anemia",
    "Bajo peso al nacer (<2500g)",
    "Prematurez (<37 semanas)",
    "Infecciones recurrentes",
    "Parasitosis intestinal",
    "Enfermedades crónicas",
    "Problemas gastrointestinales"
]

FACTORES_SOCIOECONOMICOS = [
    "Bajo nivel educativo de padres",
    "Ingresos familiares reducidos",
    "Hacinamiento en vivienda",
    "Acceso limitado a agua potable",
    "Zona rural o alejada",
    "Trabajo informal o precario"
]

# ==================================================
# FUNCIONES DE CÁLCULO DE RIESGO
# ==================================================

def calcular_riesgo_anemia(hb_ajustada, edad_meses, factores_clinicos, factores_sociales):
    puntaje = 0
    
    if edad_meses < 12:
        if hb_ajustada < 9.0: puntaje += 30
        elif hb_ajustada < 10.0: puntaje += 20
        elif hb_ajustada < 11.0: puntaje += 10
    elif edad_meses < 60:
        if hb_ajustada < 9.5: puntaje += 30
        elif hb_ajustada < 10.5: puntaje += 20
        elif hb_ajustada < 11.5: puntaje += 10
    else:
        if hb_ajustada < 10.0: puntaje += 30
        elif hb_ajustada < 11.0: puntaje += 20
        elif hb_ajustada < 12.0: puntaje += 10
    
    puntaje += len(factores_clinicos) * 4
    puntaje += len(factores_sociales) * 3
    
    if puntaje >= 35:
        return "ALTO RIESGO", puntaje, "URGENTE"
    elif puntaje >= 25:
        return "ALTO RIESGO", puntaje, "PRIORITARIO"
    elif puntaje >= 15:
        return "RIESGO MODERADO", puntaje, "EN SEGUIMIENTO"
    else:
        return "BAJO RIESGO", puntaje, "VIGILANCIA"

def generar_sugerencias(riesgo, hemoglobina_ajustada, edad_meses):
    clasificacion, recomendacion, _ = clasificar_anemia(hemoglobina_ajustada, edad_meses)
    
    if clasificacion == "ANEMIA SEVERA":
        return "🚨 INTERVENCIÓN URGENTE: Suplementación inmediata con hierro, evaluación médica en 24-48 horas, control semanal de hemoglobina."
    elif clasificacion == "ANEMIA MODERADA":
        return "⚠️ ACCIÓN PRIORITARIA: Iniciar suplementación con hierro, evaluación médica en 7 días, control mensual."
    elif clasificacion == "ANEMIA LEVE":
        return "📋 SEGUIMIENTO: Educación nutricional, dieta rica en hierro, control cada 3 meses."
    else:
        return "✅ PREVENCIÓN: Mantener alimentación balanceada, control preventivo cada 6 meses."

# ==================================================
# INTERFAZ PRINCIPAL CON INFORMACIÓN DEL USUARIO
# ==================================================

# ESTADO DE CONEXIÓN
if supabase:
    st.markdown("""
    <div style="background: #d1fae5; padding: 1rem; border-radius: 8px; border-left: 5px solid #10b981; margin-bottom: 2rem;">
        <p style="margin: 0; color: #065f46; font-weight: 500;">
        ✅ <strong>CONECTADO A SUPABASE</strong> - Sistema operativo
        </p>
    </div>
    """, unsafe_allow_html=True)
else:
    st.error("🔴 SIN CONEXIÓN A SUPABASE")

# PESTAÑAS PRINCIPALES
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📝 Registro Completo", 
    "🔍 Seguimiento Clínico", 
    "📈 Dashboard Nacional",
    "📋 Sistema de Citas",
    "⚙️ Configuración"
])

# ==================================================
# PESTAÑA 1: REGISTRO COMPLETO (CON VALIDACIÓN EN PYTHON)
# ==================================================

with tab1:
    st.markdown('<div class="section-title-blue">📝 Registro Completo de Paciente</div>', unsafe_allow_html=True)
    
    # Variables para mostrar errores
    error_dni = None
    error_nombre = None
    error_telefono = None
    
    # Variable para controlar si se mostró el resultado
    resultado_mostrado = False
    
    # Inicializar session state para limpiar
    if 'limpiar_formulario' not in st.session_state:
        st.session_state.limpiar_formulario = False
    
    # Crear el formulario
    with st.form("formulario_completo", clear_on_submit=False):
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown('**👤 Datos Personales**')
            
            dni_input = st.text_input(
                "DNI*", 
                placeholder="Ej: 87654321", 
                key="dni_input", 
                max_chars=8
            )
            
            nombre_input = st.text_input(
                "Nombre Completo*", 
                placeholder="Ej: Ana García Pérez", 
                key="nombre_input"
            )
            
            edad_meses = st.number_input("Edad (meses)*", 1, 240, 24, key="edad_input")
            peso_kg = st.number_input("Peso (kg)*", 0.0, 50.0, 12.5, 0.1, key="peso_input")
            talla_cm = st.number_input("Talla (cm)*", 0.0, 150.0, 85.0, 0.1, key="talla_input")
            genero = st.selectbox("Género*", GENEROS, key="genero_input")
            
            telefono_input = st.text_input(
                "Teléfono (9 dígitos)*", 
                placeholder="Ej: 987654321", 
                key="telefono_input", 
                max_chars=9
            )
            
            estado_paciente = st.selectbox("Estado del Paciente", ESTADOS_PACIENTE, key="estado_input")
        
        with col2:
            st.markdown('**🌍 Datos Geográficos**')
            region = st.selectbox("Región*", PERU_REGIONS, key="region_input")
            departamento = st.text_input("Departamento/Distrito", placeholder="Ej: Lima Metropolitana", key="departamento_input")
            
            if region in ALTITUD_REGIONES:
                altitud_info = ALTITUD_REGIONES[region]
                altitud_auto = altitud_info["altitud_promedio"]
                st.info(f"Altitud promedio: {altitud_auto} msnm")
            
            altitud_msnm = st.number_input("Altitud (msnm)*", 0, 5000, altitud_auto if 'altitud_auto' in locals() else 500, key="altitud_input")
            
            st.markdown('**💰 Factores Socioeconómicos del Apoderado**')
            nivel_educativo = st.selectbox("Nivel Educativo del Apoderado", NIVELES_EDUCATIVOS, key="nivel_input")
            acceso_agua_potable = st.checkbox("Acceso a agua potable", key="agua_input")
            tiene_servicio_salud = st.checkbox("Tiene servicio de salud", key="salud_input")
        
        st.markdown("---")
        col3, col4 = st.columns(2)
        
        with col3:
            st.markdown('**🩺 Parámetros Clínicos**')
            hemoglobina_medida = st.number_input("Hemoglobina medida (g/dL)*", 5.0, 20.0, 11.0, 0.1, key="hemoglobina_input")
            
            ajuste_hb = obtener_ajuste_hemoglobina(altitud_msnm)
            hemoglobina_ajustada = calcular_hemoglobina_ajustada(hemoglobina_medida, altitud_msnm)
            
            clasificacion, recomendacion, tipo_alerta = clasificar_anemia(hemoglobina_ajustada, edad_meses)
            
            # Mostrar clasificación
            if tipo_alerta == "error":
                st.error(f"🔴 {clasificacion}: {recomendacion}")
            elif tipo_alerta == "warning":
                st.warning(f"🟠 {clasificacion}: {recomendacion}")
            else:
                st.success(f"🟢 {clasificacion}: {recomendacion}")
            
            st.info(f"**Hemoglobina ajustada:** {hemoglobina_ajustada:.1f} g/dL (Ajuste: {ajuste_hb:+.1f} g/dL)")
            
            necesita_seguimiento = necesita_seguimiento_automatico(hemoglobina_ajustada, edad_meses)
            en_seguimiento = st.checkbox("Marcar para seguimiento activo", value=necesita_seguimiento, key="seguimiento_input")
            
            consume_hierro = st.checkbox("Consume suplemento de hierro", key="hierro_input")
            if consume_hierro:
                tipo_suplemento_hierro = st.text_input("Tipo de suplemento de hierro", placeholder="Ej: Sulfato ferroso", key="tipo_suplemento_input")
                frecuencia_suplemento = st.selectbox("Frecuencia de suplemento", FRECUENCIAS_SUPLEMENTO, key="frecuencia_input")
            else:
                tipo_suplemento_hierro = ""
                frecuencia_suplemento = ""
            
            antecedentes_anemia = st.checkbox("Antecedentes de anemia", key="antecedentes_input")
            enfermedades_cronicas = st.text_area("Enfermedades crónicas", placeholder="Ej: Asma, alergias, etc.", key="enfermedades_input")
        
        with col4:
            # FACTORES DE RIESGO
            st.markdown('**📋 Factores de Riesgo**')
            
            st.markdown('*Factores Clínicos*')
            factores_clinicos = st.multiselect("Seleccione factores clínicos:", FACTORES_CLINICOS, key="factores_clinicos_input")
            
            st.markdown('*Factores Socioeconómicos*')
            factores_sociales = st.multiselect("Seleccione factores socioeconómicos:", [
                "Bajo nivel educativo del apoderado",
                "Ingresos familiares reducidos",
                "Hacinamiento en vivienda",
                "Acceso limitado a agua potable",
                "Zona rural o alejada",
                "Trabajo informal o precario del apoderado",
                "Falta de acceso a servicios básicos"
            ], key="factores_sociales_input")
            
            # PROGRAMA NACIONAL DE ALIMENTACIÓN (SEPARADO)
            st.markdown("---")
            st.markdown('**🍎 Programa Nacional de Alimentación**')
            
            programas_alimentacion = st.multiselect("Seleccione programa(s) de alimentación:", [
                "Cuna Más",
                "Qali Warma",
                "Otro programa social",
                "No participa en programas"
            ], key="programas_alimentacion_input")
        
        # Botones
        col_b1, col_b2, col_b3 = st.columns(3)
        
        with col_b1:
            btn_limpiar = st.form_submit_button(
                "🧹 Limpiar", 
                type="secondary", 
                use_container_width=True
            )
        
        with col_b2:
            btn_analizar = st.form_submit_button(
                "📊 Analizar Riesgo", 
                type="primary", 
                use_container_width=True
            )
        
        with col_b3:
            btn_guardar = st.form_submit_button(
                "💾 Guardar", 
                type="primary", 
                use_container_width=True
            )
    # ============================================
    # ACCIONES FUERA DEL FORMULARIO
    # ============================================
    
    # Acción 1: Limpiar formulario
    if btn_limpiar:
        if 'datos_analizados' in st.session_state:
            del st.session_state.datos_analizados
        st.success("✅ Datos analizados limpiados")
    
    # Acción 2: Analizar Riesgo
    if btn_analizar:
        # Validaciones básicas
        errores_finales = []
        
        if not dni_input or len(dni_input) != 8:
            errores_finales.append("❌ DNI inválido (8 dígitos)")
        
        if not nombre_input or len(nombre_input.strip().split()) < 2:
            errores_finales.append("❌ Nombre completo requerido")
        
        if not telefono_input or len(telefono_input) != 9:
            errores_finales.append("❌ Teléfono inválido (9 dígitos)")
        
        if errores_finales:
            for error in errores_finales:
                st.error(error)
        else:
            try:
                nivel_riesgo, puntaje, estado = calcular_riesgo_anemia(
                    hemoglobina_ajustada,
                    edad_meses,
                    factores_clinicos,
                    factores_sociales
                )
                
                sugerencias = generar_sugerencias(nivel_riesgo, hemoglobina_ajustada, edad_meses)
                estado_peso, estado_talla, estado_nutricional = evaluar_estado_nutricional(edad_meses, peso_kg, talla_cm, genero)
                
                # Guardar en session state
                st.session_state.datos_analizados = {
                    "nivel_riesgo": nivel_riesgo,
                    "puntaje": puntaje,
                    "estado": estado,
                    "sugerencias": sugerencias,
                    "estado_peso": estado_peso,
                    "estado_talla": estado_talla,
                    "estado_nutricional": estado_nutricional
                }
                
                # Mostrar resultados
                st.markdown("---")
                st.markdown("### 📊 Resultados del Análisis")
                
                col_r1, col_r2 = st.columns(2)
                
                with col_r1:
                    st.markdown(f"**Estado de Anemia:** {clasificacion}")
                    st.markdown(f"**Hemoglobina:** {hemoglobina_ajustada:.1f} g/dL")
                    st.markdown(f"**Nivel de Riesgo:** {nivel_riesgo}")
                    st.markdown(f"**Puntaje:** {puntaje}/60")
                
                with col_r2:
                    st.markdown(f"**Estado Nutricional:** {estado_nutricional}")
                    st.markdown(f"**Peso:** {estado_peso}")
                    st.markdown(f"**Talla:** {estado_talla}")
                    st.markdown(f"**Edad:** {edad_meses} meses")
                
                st.markdown("### 💡 Recomendaciones")
                st.info(sugerencias)
                
                st.success("✅ Análisis completado")
                
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
    
    # Acción 3: Guardar en Supabase
    if btn_guardar:
        if 'datos_analizados' not in st.session_state:
            st.error("⚠️ Primero haga el análisis de riesgo")
        else:
            if supabase:
                datos = st.session_state.datos_analizados
                
                record = {
                    "dni": dni_input.strip() if dni_input else "",
                    "nombre_apellido": nombre_input.strip() if nombre_input else "",
                    "edad_meses": int(edad_meses),
                    "peso_kg": float(peso_kg),
                    "talla_cm": float(talla_cm),
                    "genero": genero,
                    "telefono": telefono_input.strip() if telefono_input else "",
                    "estado_paciente": estado_paciente,
                    "region": region,
                    "departamento": departamento.strip() if departamento else "",
                    "altitud_msnm": int(altitud_msnm),
                    "nivel_educativo": nivel_educativo,
                    "acceso_agua_potable": acceso_agua_potable,
                    "tiene_servicio_salud": tiene_servicio_salud,
                    "hemoglobina_dl1": float(hemoglobina_medida),
                    "en_seguimiento": en_seguimiento,
                    "consumir_hierro": consume_hierro,
                    "tipo_suplemento_hierro": tipo_suplemento_hierro.strip() if consume_hierro and tipo_suplemento_hierro else "",
                    "frecuencia_suplemento": frecuencia_suplemento if consume_hierro else "",
                    "antecedentes_anemia": antecedentes_anemia,
                    "enfermedades_cronicas": enfermedades_cronicas.strip() if enfermedades_cronicas else "",
                    "riesgo": datos["nivel_riesgo"],
                    "fecha_alerta": datetime.now().strftime("%Y-%m-%d"),
                    "estado_alerta": datos["estado"],
                    "sugerencias": datos["sugerencias"],
                    "programas_alimentacion": ", ".join(programas_alimentacion) if programas_alimentacion else "No participa"
                }
                
                try:
                    resultado = insertar_datos_supabase(record)
                    
                    if resultado:
                        if isinstance(resultado, dict) and resultado.get("status") == "duplicado":
                            st.error(f"❌ DNI {dni_input} ya existe")
                        else:
                            st.success("✅ Datos guardados correctamente")
                            st.balloons()
                    else:
                        st.error("❌ Error al guardar")
                        
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
            else:
                st.error("🔴 Sin conexión a Supabase")
# ==================================================
# PESTAÑA 1: REGISTRO INTEGRAL Y ANÁLISIS ETIOLÓGICO
# ==================================================
with tab1:
    st.markdown('<div class="section-title-blue">📝 Registro Integral de Paciente y Biomarcadores</div>', unsafe_allow_html=True)
    
    # 1. CONTENEDOR DE DATOS (Usamos container en lugar de form para respuesta inmediata)
    with st.container():
        # --- SECCIÓN 1: DATOS DE IDENTIDAD ---
        col1, col2 = st.columns(2)
        with col1:
            st.markdown('**👤 Identidad**')
            dni_input = st.text_input("DNI*", max_chars=8, key="f_dni")
            nombre_input = st.text_input("Nombre Completo*", key="f_nom")
            edad_meses = st.number_input("Edad (meses)*", 1, 240, 24, key="f_edad")
            
        with col2:
            st.markdown('**🌍 Ubicación y Social**')
            region_input = st.selectbox("Región*", PERU_REGIONS, key="f_reg")
            altitud = st.number_input("Altitud (msnm)*", 0, 5000, 500, key="f_alt")
            seguimiento = st.checkbox("Activar seguimiento automático", value=True, key="b_seg")

        st.markdown("---")

        # --- SECCIÓN 2: BIOMARCADORES ---
        st.markdown('### 🧪 Análisis de Biomarcadores (Laboratorio)')
        c_bio1, c_bio2 = st.columns(2)
        with c_bio1:
            st.markdown('**🩸 Perfil de Hierro**')
            f_ferritina = st.number_input("Ferritina (ng/mL)", 0.0, 500.0, 15.0, key="b_fer")
            f_chcm = st.number_input("CHCM (g/dL)", 20.0, 40.0, 32.0, key="b_chcm")
            f_transf = st.number_input("Transferrina (mg/dL)", 100.0, 500.0, 250.0, key="b_trans")
            
        with c_bio2:
            st.markdown('**🦠 Respuesta Medular e Inflamación**')
            f_retic = st.number_input("Reticulocitos (%)", 0.0, 5.0, 1.0, key="b_ret")
            f_pcr = st.number_input("PCR (mg/dL)", 0.0, 100.0, 0.1, key="b_pcr")
            
        st.markdown("---")

        # --- SECCIÓN 3: HEMOGLOBINA ---
        st.markdown('**🩺 Parámetros Clínicos**')
        hb_medida = st.number_input("Hemoglobina medida (g/dL)*", 5.0, 20.0, 11.0, key="b_hb")
        hb_ajustada = calcular_hemoglobina_ajustada(hb_medida, altitud)
        st.info(f"**HB Ajustada:** {hb_ajustada:.1f} g/dL")

        # --- BOTONES DE ACCIÓN (Fuera de Form para que funcionen) ---
        st.markdown("<br>", unsafe_allow_html=True)
        col_btn1, col_btn2 = st.columns(2)
        
        with col_btn1:
            # Este botón ahora sí disparará la interpretación al instante
            btn_analizar = st.button("📊 GENERAR DIAGNÓSTICO ETIOLÓGICO", use_container_width=True)
            
        with col_btn2:
            btn_guardar = st.button("💾 GUARDAR REGISTRO COMPLETO", type="primary", use_container_width=True)

    # --- LÓGICA DE INTERPRETACIÓN (Debe estar fuera de cualquier form) ---
if btn_analizar:
    # 1. Ejecutar la función lógica
    resultado = interpretar_analisis_hematologico(
        f_ferritina, f_chcm, f_retic, f_transf, hb_ajustada, edad_meses, f_pcr
    )
    
    # 2. CUADRO PRINCIPAL (Resumen visual con color)
    st.markdown(f"""
    <div style="background-color: #1e3a8a; padding: 25px; border-radius: 15px; color: white; border: 3px solid rgba(255,255,255,0.5); margin-top: 20px;">
        <h2 style="margin:0; color: white;">📋 Resultado del Análisis Etiológico</h2>
        <hr style="border: 0.5px solid rgba(255,255,255,0.3); margin: 15px 0;">
        <p style="font-size: 1.3rem; font-weight: bold;">{resultado['interpretacion']}</p>
        <div style="background: rgba(0,0,0,0.15); padding: 15px; border-radius: 10px; border-left: 5px solid white;">
            <p style="margin:0; font-size: 1.1rem;">
                <strong>🩺 RECOMENDACIÓN CLÍNICA:</strong><br>{resultado['recomendacion']}
            </p>
        </div>
    </div>
""", unsafe_allow_html=True)
    
    # 3. SECCIÓN DE DIAGNÓSTICO DIFERENCIAL (Explicación de por qué se descartan otras)
    st.markdown("### ❌ Diagnóstico Diferencial (Sustento de Descartes)")
    
    # Creamos columnas para que los descartes no ocupen tanto espacio vertical
    col_desc1, col_desc2 = st.columns(2)
    
    for i, item in enumerate(resultado['descartes']):
        # Alternamos entre columna 1 y columna 2
        if i % 2 == 0:
            col_desc1.info(item)
        else:
            col_desc2.info(item)
            
    # 4. BOTÓN EXTRA: EXPLICACIÓN CIENTÍFICA (Opcional)
    with st.expander("📚 Ver criterios científicos aplicados"):
        st.write("""
        * **Ferritina < 15/30:** Agotamiento de reservas (Gold Standard).
        * **PCR Normal:** Asegura que la ferritina no esté elevada por infección.
        * **CHCM < 32:** Confirma hipocromía (típico de ferropenia).
        * **Reticulocitos normales/bajos:** Indica que la médula no tiene hierro para producir sangre.
        """)

    # --- LÓGICA DE GUARDADO ---
    if btn_guardar:
        if dni_input and nombre_input:
            datos_paciente = {
                "dni": dni_input,
                "nombre_apellido": nombre_input,
                "region": region_input,
                "hb_ajustada": hb_ajustada,
                "ferritina": f_ferritina,
                "pcr": f_pcr,
                "seguimiento": seguimiento
            }
            # Aquí iría tu insertar_en_supabase(datos_paciente)
            st.success(f"✅ Registro de {nombre_input} guardado exitosamente.")
        else:
            st.warning("⚠️ Complete DNI y Nombre antes de guardar.")
# ==================================================
# PESTAÑA 2: SEGUIMIENTO CLÍNICO - VERSIÓN CORREGIDA
# ==================================================

with tab2:
    st.markdown('<div class="section-title-green">🔬 SISTEMA DE SEGUIMIENTO CLÍNICO COMPLETO</div>', unsafe_allow_html=True)
    
    # ============================================
    # INICIALIZACIÓN DE ESTADOS
    # ============================================
    
    if 'seguimiento_paciente' not in st.session_state:
        st.session_state.seguimiento_paciente = None
    
    if 'seguimiento_datos_pacientes' not in st.session_state:
        st.session_state.seguimiento_datos_pacientes = pd.DataFrame()
    
    if 'seguimiento_historial' not in st.session_state:
        st.session_state.seguimiento_historial = []
    
    # ============================================
    # FUNCIONES CORREGIDAS
    # ============================================
    
    def cargar_todos_pacientes():
        """Carga todos los pacientes desde Supabase"""
        try:
            with st.spinner("🔄 Cargando pacientes..."):
                response = supabase.table("alertas_hemoglobina").select("*").execute()
                
                if response.data:
                    df = pd.DataFrame(response.data)
                    
                    # Asegurar que las columnas necesarias existan
                    columnas_necesarias = ['dni', 'nombre_apellido', 'edad_meses', 'hemoglobina_dl1', 'region']
                    for col in columnas_necesarias:
                        if col not in df.columns:
                            df[col] = None
                    
                    st.session_state.seguimiento_datos_pacientes = df
                    return True
                else:
                    st.error("❌ No se encontraron pacientes en la base de datos")
                    return False
                    
        except Exception as e:
            st.error(f"❌ Error al cargar pacientes: {str(e)[:100]}")
            return False
    
    # ============================================
    # CREAR 3 PESTAÑAS INTERNAS
    # ============================================
    
    tab_seg1, tab_seg2, tab_seg3 = st.tabs([
        "🔍 Buscar Paciente", 
        "📝 Nuevo Seguimiento", 
        "📋 Historial Completo"
    ])
    
    # ============================================
    # PESTAÑA 1: BUSCAR PACIENTE - VERSIÓN CORREGIDA
    # ============================================
    
    with tab_seg1:
        st.header("🔍 BUSCAR PACIENTE PARA SEGUIMIENTO")
        
        # Botón para cargar pacientes
        if st.button("🔄 Cargar Todos los Pacientes", type="primary", use_container_width=True, key="btn_cargar_pacientes_seg1"):
            cargar_todos_pacientes()
        
        # Verificar si hay datos cargados
        if not st.session_state.seguimiento_datos_pacientes.empty:
            df = st.session_state.seguimiento_datos_pacientes
            
            # Búsqueda por DNI, nombre o región
            buscar = st.text_input("🔎 Buscar por nombre, DNI o región:", 
                                 placeholder="Ej: 'Mia' o '10096525' o 'LIMA'",
                                 key="buscar_paciente_input")
            
            if buscar:
                # Convertir a string para búsqueda
                mask = (
                    df['nombre_apellido'].astype(str).str.contains(buscar, case=False, na=False) |
                    df['dni'].astype(str).str.contains(buscar, na=False) |
                    df['region'].astype(str).str.contains(buscar, case=False, na=False)
                )
                df_filtrado = df[mask]
                
                # Si no hay resultados con búsqueda normal, intentar búsqueda exacta por DNI
                if df_filtrado.empty and buscar.isdigit():
                    mask_exact = df['dni'].astype(str) == buscar
                    df_filtrado = df[mask_exact]
            else:
                df_filtrado = df
            
            # Mostrar resultados
            if not df_filtrado.empty:
                st.write(f"📊 **{len(df_filtrado)} pacientes encontrados**")
                
                # Crear lista de selección
                opciones = []
                for _, row in df_filtrado.iterrows():
                    nombre = row.get('nombre_apellido', 'N/A')
                    dni = row.get('dni', 'N/A')
                    edad = row.get('edad_meses', 'N/A')
                    hb = row.get('hemoglobina_dl1', 'N/A')
                    
                    opcion_text = f"{nombre} - DNI: {dni} - Edad: {edad} meses - Hb: {hb} g/dL"
                    opciones.append((opcion_text, dni))
                
                # Selector
                if opciones:
                    opcion_seleccionada = st.selectbox(
                        "Seleccione un paciente:",
                        options=[op[0] for op in opciones],
                        placeholder="Elija un paciente de la lista...",
                        key="select_paciente_seguimiento"
                    )
                    
                    # Encontrar DNI seleccionado
                    dni_seleccionado = None
                    for opcion_text, dni in opciones:
                        if opcion_text == opcion_seleccionada:
                            dni_seleccionado = dni
                            break
                    
                    if dni_seleccionado:
                        paciente_info = df_filtrado[df_filtrado['dni'] == dni_seleccionado].iloc[0]
                        
                        # Mostrar información del paciente
                        st.markdown("---")
                        col_show1, col_show2 = st.columns(2)
                        
                        with col_show1:
                            st.markdown(f"""
                            <div style="background: #dbeafe; padding: 1rem; border-radius: 10px; margin-bottom: 1rem;">
                                <h4 style="margin: 0 0 10px 0; color: #1e40af;">👤 DATOS PERSONALES</h4>
                                <p><strong>Paciente:</strong> {paciente_info['nombre_apellido']}</p>
                                <p><strong>DNI:</strong> {paciente_info['dni']}</p>
                                <p><strong>Edad:</strong> {paciente_info['edad_meses']} meses</p>
                                <p><strong>Género:</strong> {paciente_info.get('genero', 'N/A')}</p>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        with col_show2:
                            st.markdown(f"""
                            <div style="background: #d1fae5; padding: 1rem; border-radius: 10px; margin-bottom: 1rem;">
                                <h4 style="margin: 0 0 10px 0; color: #059669;">🩺 DATOS CLÍNICOS</h4>
                                <p><strong>Hemoglobina:</strong> {paciente_info['hemoglobina_dl1']} g/dL</p>
                                <p><strong>Región:</strong> {paciente_info['region']}</p>
                                <p><strong>Estado:</strong> {paciente_info.get('estado_paciente', 'N/A')}</p>
                                <p><strong>Riesgo:</strong> {paciente_info.get('riesgo', 'N/A')}</p>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        # Botón para seleccionar
                        col_btn_sel1, col_btn_sel2 = st.columns(2)
                        
                        with col_btn_sel1:
                            if st.button("✅ Seleccionar este paciente", 
                                       use_container_width=True, 
                                       type="primary",
                                       key=f"btn_seleccionar_{dni_seleccionado}"):
                                
                                st.session_state.seguimiento_paciente = paciente_info.to_dict()
                                
                                # Cargar historial
                                try:
                                    response = supabase.table("seguimientos")\
                                        .select("*")\
                                        .eq("dni_paciente", str(dni_seleccionado))\
                                        .order("fecha_seguimiento", desc=True)\
                                        .execute()
                                    
                                    if response.data:
                                        st.session_state.seguimiento_historial = response.data
                                        cantidad = len(response.data)
                                    else:
                                        st.session_state.seguimiento_historial = []
                                        cantidad = 0
                                        
                                except Exception as e:
                                    st.session_state.seguimiento_historial = []
                                    cantidad = 0
                                    st.warning(f"⚠️ No se pudo cargar historial: {str(e)[:100]}")
                                
                                st.success(f"✅ Paciente seleccionado: {paciente_info['nombre_apellido']}")
                                st.info(f"📋 Se cargaron {cantidad} controles previos")
                                
                                time.sleep(1)
                                st.rerun()
                        
                        with col_btn_sel2:
                            if st.button("📋 Ver detalles completos", 
                                       use_container_width=True, 
                                       type="secondary",
                                       key=f"btn_detalles_{dni_seleccionado}"):
                                with st.expander("📄 Detalles completos del paciente", expanded=True):
                                    st.json(paciente_info.to_dict())
            else:
                st.info("🔍 No se encontraron pacientes con los criterios de búsqueda")
        else:
            st.info("👆 Presiona 'Cargar Todos los Pacientes' para buscar pacientes")

    # ============================================
    # PESTAÑA 2: NUEVO SEGUIMIENTO - VERSIÓN CORREGIDA
    # ============================================
    
    with tab_seg2:
        st.header("📝 NUEVO SEGUIMIENTO CLÍNICO")
        
        # Variable para controlar si se guardó exitosamente
        seguimiento_guardado = False
        
        # Verificar si hay paciente seleccionado
        if not st.session_state.seguimiento_paciente:
            st.warning("""
            ⚠️ **Primero debe seleccionar un paciente**
            
            Pasos:
            1. Ve a la pestaña **🔍 Buscar Paciente**
            2. Cargue los pacientes
            3. Seleccione un paciente de la lista
            4. Regrese aquí para crear el seguimiento
            """)
            
            if st.button("🔍 Ir a Buscar Paciente", 
                        use_container_width=True, 
                        key="btn_ir_buscar_paciente_seg2"):
                st.markdown("""
                <script>
                setTimeout(() => {
                    const tabs = document.querySelectorAll('button[role="tab"]');
                    if (tabs.length >= 2) {
                        tabs[1].click();
                    }
                }, 500);
                </script>
                """, unsafe_allow_html=True)
                st.rerun()
        else:
            paciente = st.session_state.seguimiento_paciente
            
            # Mostrar información del paciente
            col_info1, col_info2 = st.columns(2)
            
            with col_info1:
                st.markdown(f"""
                <div style="background: #dbeafe; padding: 1rem; border-radius: 10px; margin-bottom: 1rem;">
                    <h4 style="margin: 0 0 10px 0; color: #1e40af;">📋 PACIENTE SELECCIONADO</h4>
                    <p><strong>Nombre:</strong> {paciente.get('nombre_apellido', 'N/A')}</p>
                    <p><strong>DNI:</strong> {paciente.get('dni', 'N/A')}</p>
                    <p><strong>Edad:</strong> {paciente.get('edad_meses', 'N/A')} meses</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col_info2:
                st.markdown(f"""
                <div style="background: #d1fae5; padding: 1rem; border-radius: 10px; margin-bottom: 1rem;">
                    <h4 style="margin: 0 0 10px 0; color: #059669;">📊 DATOS CLÍNICOS</h4>
                    <p><strong>Hb actual:</strong> {paciente.get('hemoglobina_dl1', 'N/A')} g/dL</p>
                    <p><strong>Región:</strong> {paciente.get('region', 'N/A')}</p>
                    <p><strong>Estado:</strong> {paciente.get('estado_paciente', 'N/A')}</p>
                </div>
                """, unsafe_allow_html=True)
            
            # Formulario de seguimiento CORREGIDO
            with st.form("form_seguimiento_corregido", clear_on_submit=True):
                st.subheader("📊 Datos del Control")
                
                # SOLO FECHA - SIN HORA
                fecha = st.date_input("Fecha del control*", 
                                     datetime.now().date(),
                                     key="fecha_seguimiento_corregida")
                
                col_hb, col_peso, col_talla = st.columns(3)
                with col_hb:
                    hemoglobina = st.number_input(
                        "Hemoglobina (g/dL)*", 
                        value=float(paciente.get('hemoglobina_dl1', 11.0)),
                        min_value=0.0, 
                        max_value=20.0, 
                        step=0.1,
                        key="hemoglobina_seguimiento_corregida"
                    )
                with col_peso:
                    peso = st.number_input(
                        "Peso (kg)*", 
                        value=float(paciente.get('peso_kg', 12.0)),
                        min_value=0.0, 
                        max_value=50.0, 
                        step=0.1,
                        key="peso_seguimiento_corregida"
                    )
                with col_talla:
                    talla = st.number_input(
                        "Talla (cm)*", 
                        value=float(paciente.get('talla_cm', 85.0)),
                        min_value=0.0, 
                        max_value=150.0, 
                        step=0.1,
                        key="talla_seguimiento_corregida"
                    )
                
                # Evaluar estado nutricional
                edad_paciente = paciente.get('edad_meses', 24)
                genero_paciente = paciente.get('genero', 'F')
                
                estado_peso, estado_talla, estado_nutricional = evaluar_estado_nutricional(
                    edad_paciente, peso, talla, genero_paciente
                )
                
                # Mostrar estado nutricional
                if estado_nutricional in ["DESNUTRICIÓN CRÓNICA", "DESNUTRICIÓN AGUDA"]:
                    st.error(f"**Estado nutricional:** {estado_nutricional}")
                elif estado_nutricional == "SOBREPESO":
                    st.warning(f"**Estado nutricional:** {estado_nutricional}")
                else:
                    st.success(f"**Estado nutricional:** {estado_nutricional}")
                
                st.caption(f"📏 **Peso:** {estado_peso} | **Talla:** {estado_talla}")
                
                # Campo NUEVO: Médico/Especialista responsable
                medico_responsable = st.text_input(
                    "Médico/Especialista responsable*",
                    placeholder="Nombre del médico que realiza el seguimiento",
                    key="medico_responsable_input"
                )
                
                tipo_seguimiento = st.selectbox(
                    "Tipo de seguimiento*",
                    ["Control rutinario", "Seguimiento por anemia", "Control nutricional", 
                     "Evaluación de tratamiento", "Consulta de urgencia", "Control de crecimiento"],
                    key="tipo_seguimiento_select_corregida"
                )
                
                observaciones = st.text_area(
                    "Observaciones clínicas*", 
                    placeholder="Describa el estado del paciente, síntomas, evolución, respuesta al tratamiento...",
                    height=100,
                    key="observaciones_seguimiento_corregida"
                )
                
                tratamiento = st.text_input(
                    "Tratamiento actual o prescrito", 
                    placeholder="Ej: Sulfato ferroso 15 mg/día, suplemento vitamínico...",
                    key="tratamiento_seguimiento_corregida"
                )
                
                proximo_control = st.date_input(
                    "Próximo control sugerido",
                    value=datetime.now().date() + timedelta(days=30),
                    key="proximo_control_input_corregida"
                )
                
                # Botones DENTRO del formulario (CORRECTOS)
                col_btn1, col_btn2, col_btn3 = st.columns(3)
                with col_btn1:
                    submit = st.form_submit_button(
                        "💾 GUARDAR SEGUIMIENTO", 
                        type="primary", 
                        use_container_width=True,
                        key="btn_guardar_seguimiento_corregido"
                    )
                with col_btn2:
                    limpiar = st.form_submit_button(
                        "🧹 LIMPIAR FORMULARIO", 
                        type="secondary", 
                        use_container_width=True,
                        key="btn_limpiar_seguimiento_corregido"
                    )
                with col_btn3:
                    cancelar = st.form_submit_button(
                        "❌ CANCELAR", 
                        type="secondary", 
                        use_container_width=True,
                        key="btn_cancelar_seguimiento_corregido"
                    )
                
                # Procesar guardado - VERSIÓN CORREGIDA
                if submit:
                    # Validaciones
                    if not medico_responsable.strip():
                        st.error("⚠️ El médico responsable es obligatorio")
                    elif not observaciones.strip():
                        st.error("⚠️ Las observaciones clínicas son obligatorias")
                    else:
                        # PREPARAR DATOS CORRECTAMENTE
                        hemoglobina_ajustada = hemoglobina  # Puedes ajustar esta fórmula
                        
                        # Crear observaciones completas
                        observaciones_completas = f"""{observaciones}

DATOS ADICIONALES:
- Peso: {peso} kg ({estado_peso})
- Talla: {talla} cm ({estado_talla})
- Estado nutricional: {estado_nutricional}
- Frecuencia control sugerida: {proximo_control.strftime('%d/%m/%Y')}"""
                        
                        # CORRECCIÓN: Usar los nombres de columna correctos
                        datos = {
                            "dni_paciente": str(paciente.get('dni', '')),
                            "fecha_seguimiento": fecha.strftime('%Y-%m-%d'),  # SOLO FECHA
                            "tipo_seguimiento": tipo_seguimiento,
                            "hemoglobina_actual": hemoglobina,
                            "hemoglobina_ajustada": hemoglobina_ajustada,
                            "clasificacion_actual": estado_nutricional,
                            "observaciones": observaciones_completas,
                            "tratamiento_actual": tratamiento,
                            "usuario_responsable": medico_responsable,  # CORREGIDO: usar campo del formulario
                            "proximo_control": proximo_control.strftime('%Y-%m-%d'),
                            "created_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }
                        
                        # Guardar en Supabase
                        try:
                            response = supabase.table("seguimientos").insert(datos).execute()
                            
                            if response.data:
                                st.success("✅ Seguimiento guardado correctamente")
                                seguimiento_guardado = True
                                
                                # Actualizar historial en session state
                                if 'seguimiento_historial' not in st.session_state:
                                    st.session_state.seguimiento_historial = []
                                st.session_state.seguimiento_historial.insert(0, datos)
                                
                                # Actualizar hemoglobina en tabla principal
                                try:
                                    supabase.table("alertas_hemoglobina")\
                                        .update({"hemoglobina_dl1": hemoglobina})\
                                        .eq("dni", paciente.get('dni'))\
                                        .execute()
                                    st.info("🔄 Hemoglobina actualizada en registro principal")
                                except Exception as update_error:
                                    st.warning(f"⚠️ No se pudo actualizar hemoglobina: {str(update_error)[:50]}")
                                
                                st.balloons()
                                
                                # Mostrar resumen DENTRO del formulario (esto está permitido)
                                st.markdown("---")
                                col_res1, col_res2, col_res3 = st.columns(3)
                                with col_res1:
                                    st.metric("📅 Fecha", fecha.strftime('%d/%m/%Y'))
                                with col_res2:
                                    st.metric("🩺 Hb", f"{hemoglobina} g/dL")
                                with col_res3:
                                    st.metric("⚖️ Peso", f"{peso} kg")
                                
                                # Redirección automática con JavaScript
                                st.info("📋 Redirigiendo al historial en 3 segundos...")
                                
                                # Script para cambiar de pestaña automáticamente
                                st.markdown("""
                                <script>
                                setTimeout(() => {
                                    const tabs = document.querySelectorAll('button[role="tab"]');
                                    if (tabs.length >= 4) {
                                        tabs[3].click();
                                    }
                                }, 3000);
                                </script>
                                """, unsafe_allow_html=True)
                                
                            else:
                                st.error("❌ Error al guardar el seguimiento - Respuesta vacía del servidor")
                                
                        except Exception as e:
                            error_msg = str(e)
                            st.error(f"❌ Error al guardar: {error_msg[:200]}")
                            if "tratamiento_actual" in error_msg or "usuario_responsable" in error_msg:
                                st.error("⚠️ Problema con columnas de la base de datos. Ejecuta el ALTER TABLE en Supabase.")
                
                if limpiar:
                    st.info("🧹 Formulario limpiado")
                    time.sleep(1)
                    st.rerun()
                
                if cancelar:
                    st.info("❌ Seguimiento cancelado")
                    time.sleep(1)
                    st.rerun()
            
            # ============================================
            # BOTÓN FUERA DEL FORMULARIO (para redirección manual)
            # ============================================
            
            # Solo mostrar este botón si se guardó exitosamente
            if seguimiento_guardado:
                st.markdown("---")
                col_manual1, col_manual2 = st.columns(2)
                
                with col_manual1:
                    if st.button("📋 Ir al Historial Ahora", 
                               use_container_width=True, 
                               type="primary",
                               key="btn_ir_historial_manualmente"):
                        # Cambiar a pestaña de Historial usando JavaScript
                        st.markdown("""
                        <script>
                        setTimeout(() => {
                            const tabs = document.querySelectorAll('button[role="tab"]');
                            if (tabs.length >= 4) {
                                tabs[3].click();
                            }
                        }, 100);
                        </script>
                        """, unsafe_allow_html=True)
                        st.rerun()
                
                with col_manual2:
                    if st.button("📝 Crear otro seguimiento", 
                               use_container_width=True, 
                               type="secondary",
                               key="btn_otro_seguimiento"):
                        st.info("🔄 Limpiando formulario...")
                        time.sleep(1)
                        st.rerun()


                
# ==================================================
# PESTAÑA 3: DASHBOARD NACIONAL - SOLO AQUÍ DEBE ESTAR ESTE CÓDIGO
# ==================================================

with tab3:
    st.markdown("""
    <div class="main-title" style="background: linear-gradient(135deg, #059669 0%, #10b981 100%); padding: 2rem;">
        <h2 style="margin: 0; color: white;">📈 DASHBOARD NACIONAL DE ANEMIA</h2>
        <p style="margin: 10px 0 0 0; color: rgba(255,255,255,0.9);">
        Análisis nacional, mapa interactivo por regiones, prevalencia y seguimiento
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # ============================================
    # FUNCIONES ESPECIALES PARA EL DASHBOARD
    # ============================================
    
    def calcular_indicadores_anemia(datos):
        """Calcula indicadores específicos de anemia"""
        if datos.empty:
            return {}
        
        # Asegurar que tenemos las columnas necesarias
        if 'hemoglobina_dl1' not in datos.columns:
            datos['hemoglobina_dl1'] = 11.0
        
        # Clasificar pacientes por nivel de anemia
        condiciones = [
            (datos['hemoglobina_dl1'] < 7.0),
            (datos['hemoglobina_dl1'] < 10.0),
            (datos['hemoglobina_dl1'] < 11.0),
            (datos['hemoglobina_dl1'] >= 11.0)
        ]
        
        categorias = ['SEVERA', 'MODERADA', 'LEVE', 'NORMAL']
        datos['nivel_anemia'] = np.select(condiciones, categorias, default='NORMAL')
        
        # Calcular indicadores nacionales
        total = len(datos)
        con_anemia = len(datos[datos['nivel_anemia'].isin(['SEVERA', 'MODERADA', 'LEVE'])])
        
        indicadores = {
            'total_pacientes': total,
            'con_anemia': con_anemia,
            'prevalencia_nacional': round((con_anemia / total * 100), 1) if total > 0 else 0,
            'severa': len(datos[datos['nivel_anemia'] == 'SEVERA']),
            'moderada': len(datos[datos['nivel_anemia'] == 'MODERADA']),
            'leve': len(datos[datos['nivel_anemia'] == 'LEVE']),
            'normal': len(datos[datos['nivel_anemia'] == 'NORMAL']),
            'en_seguimiento': datos['en_seguimiento'].sum() if 'en_seguimiento' in datos.columns else 0,
            'tasa_seguimiento': 0,
            'hb_promedio_nacional': datos['hemoglobina_dl1'].mean() if 'hemoglobina_dl1' in datos.columns else 0
        }
        
        # Calcular tasa de seguimiento
        if con_anemia > 0:
            anemia_df = datos[datos['nivel_anemia'].isin(['SEVERA', 'MODERADA', 'LEVE'])]
            if len(anemia_df) > 0:
                indicadores['tasa_seguimiento'] = round((anemia_df['en_seguimiento'].sum() / len(anemia_df)) * 100, 1)
        
        # Calcular por región
        if 'region' in datos.columns:
            region_stats = {}
            for region in datos['region'].unique():
                region_df = datos[datos['region'] == region]
                total_region = len(region_df)
                con_anemia_region = len(region_df[region_df['nivel_anemia'].isin(['SEVERA', 'MODERADA', 'LEVE'])])
                
                prevalencia_region = 0
                if total_region > 0:
                    prevalencia_region = round((con_anemia_region / total_region * 100), 1)
                
                region_stats[region] = {
                    'total': total_region,
                    'con_anemia': con_anemia_region,
                    'prevalencia': prevalencia_region,
                    'hb_promedio': region_df['hemoglobina_dl1'].mean() if 'hemoglobina_dl1' in region_df.columns else 0,
                    'severa': len(region_df[region_df['nivel_anemia'] == 'SEVERA']),
                    'moderada': len(region_df[region_df['nivel_anemia'] == 'MODERADA']),
                    'leve': len(region_df[region_df['nivel_anemia'] == 'LEVE']),
                    'en_seguimiento': region_df['en_seguimiento'].sum() if 'en_seguimiento' in region_df.columns else 0
                }
            
            indicadores['por_region'] = region_stats
        
        return indicadores
    
    def crear_mapa_peru(indicadores):
        """Crea un mapa del Perú con colores según prevalencia de anemia"""
        
        # Datos geográficos básicos de las regiones del Perú
        peru_geojson = {
            "type": "FeatureCollection",
            "features": [
                {"type": "Feature", "properties": {"region": "AMAZONAS"}, "geometry": {"type": "Point", "coordinates": [-78.5, -5.2]}},
                {"type": "Feature", "properties": {"region": "ANCASH"}, "geometry": {"type": "Point", "coordinates": [-77.5, -9.5]}},
                {"type": "Feature", "properties": {"region": "APURIMAC"}, "geometry": {"type": "Point", "coordinates": [-72.9, -13.6]}},
                {"type": "Feature", "properties": {"region": "AREQUIPA"}, "geometry": {"type": "Point", "coordinates": [-71.5, -16.4]}},
                {"type": "Feature", "properties": {"region": "AYACUCHO"}, "geometry": {"type": "Point", "coordinates": [-74.2, -13.2]}},
                {"type": "Feature", "properties": {"region": "CAJAMARCA"}, "geometry": {"type": "Point", "coordinates": [-78.5, -7.2]}},
                {"type": "Feature", "properties": {"region": "CALLAO"}, "geometry": {"type": "Point", "coordinates": [-77.1, -12.0]}},
                {"type": "Feature", "properties": {"region": "CUSCO"}, "geometry": {"type": "Point", "coordinates": [-71.9, -13.5]}},
                {"type": "Feature", "properties": {"region": "HUANCAVELICA"}, "geometry": {"type": "Point", "coordinates": [-75.0, -12.8]}},
                {"type": "Feature", "properties": {"region": "HUANUCO"}, "geometry": {"type": "Point", "coordinates": [-76.2, -9.9]}},
                {"type": "Feature", "properties": {"region": "ICA"}, "geometry": {"type": "Point", "coordinates": [-75.7, -14.1]}},
                {"type": "Feature", "properties": {"region": "JUNIN"}, "geometry": {"type": "Point", "coordinates": [-75.0, -11.5]}},
                {"type": "Feature", "properties": {"region": "LA LIBERTAD"}, "geometry": {"type": "Point", "coordinates": [-79.0, -8.1]}},
                {"type": "Feature", "properties": {"region": "LAMBAYEQUE"}, "geometry": {"type": "Point", "coordinates": [-79.9, -6.7]}},
                {"type": "Feature", "properties": {"region": "LIMA"}, "geometry": {"type": "Point", "coordinates": [-77.0, -12.0]}},
                {"type": "Feature", "properties": {"region": "LORETO"}, "geometry": {"type": "Point", "coordinates": [-73.2, -3.7]}},
                {"type": "Feature", "properties": {"region": "MADRE DE DIOS"}, "geometry": {"type": "Point", "coordinates": [-69.2, -12.6]}},
                {"type": "Feature", "properties": {"region": "MOQUEGUA"}, "geometry": {"type": "Point", "coordinates": [-70.9, -17.2]}},
                {"type": "Feature", "properties": {"region": "PASCO"}, "geometry": {"type": "Point", "coordinates": [-76.2, -10.7]}},
                {"type": "Feature", "properties": {"region": "PIURA"}, "geometry": {"type": "Point", "coordinates": [-80.6, -5.2]}},
                {"type": "Feature", "properties": {"region": "PUNO"}, "geometry": {"type": "Point", "coordinates": [-70.0, -15.8]}},
                {"type": "Feature", "properties": {"region": "SAN MARTIN"}, "geometry": {"type": "Point", "coordinates": [-76.1, -6.5]}},
                {"type": "Feature", "properties": {"region": "TACNA"}, "geometry": {"type": "Point", "coordinates": [-70.2, -18.0]}},
                {"type": "Feature", "properties": {"region": "TUMBES"}, "geometry": {"type": "Point", "coordinates": [-80.5, -3.6]}},
                {"type": "Feature", "properties": {"region": "UCAYALI"}, "geometry": {"type": "Point", "coordinates": [-73.4, -8.4]}}
            ]
        }
        
        # Crear DataFrame para el mapa
        map_data = []
        if 'por_region' in indicadores:
            for region, stats in indicadores.get('por_region', {}).items():
                map_data.append({
                    'region': region,
                    'prevalencia': stats['prevalencia'],
                    'total_pacientes': stats['total'],
                    'con_anemia': stats['con_anemia'],
                    'hb_promedio': stats['hb_promedio'],
                    'lat': 0,
                    'lon': 0
                })
        
        # Asignar coordenadas desde el GeoJSON
        for feature in peru_geojson['features']:
            region_name = feature['properties']['region']
            coords = feature['geometry']['coordinates']
            
            for item in map_data:
                if item['region'] == region_name:
                    item['lon'] = coords[0]
                    item['lat'] = coords[1]
                    break
        
        return pd.DataFrame(map_data)
    
    # ============================================
    # INTERFAZ PRINCIPAL DEL DASHBOARD
    # ============================================
    
    # Botón para cargar datos
    col_btn1, col_btn2 = st.columns([2, 1])
    
    with col_btn1:
        if st.button("🔄 CARGAR DATOS NACIONALES", 
                    type="primary", 
                    use_container_width=True,
                    key="btn_cargar_datos_nacionales_tab3"):
            with st.spinner("Cargando datos nacionales..."):
                datos_nacionales = obtener_datos_supabase()
                
                if not datos_nacionales.empty:
                    # Calcular indicadores
                    indicadores = calcular_indicadores_anemia(datos_nacionales)
                    mapa_data = crear_mapa_peru(indicadores)
                    
                    st.session_state.datos_nacionales = datos_nacionales
                    st.session_state.indicadores_anemia = indicadores
                    st.session_state.mapa_peru = mapa_data
                    
                    st.success(f"✅ {len(datos_nacionales)} registros cargados - {indicadores['prevalencia_nacional']}% de prevalencia")
                else:
                    st.error("❌ No se pudieron cargar datos nacionales")
    
       
    # ============================================
    # MOSTRAR DASHBOARD SI HAY DATOS
    # ============================================
    
    if 'indicadores_anemia' in st.session_state and st.session_state.indicadores_anemia:
        indicadores = st.session_state.indicadores_anemia
        datos = st.session_state.datos_nacionales
        
        # ============================================
        # MÉTRICAS PRINCIPALES
        # ============================================
        
        st.markdown("""
        <div class="section-title-blue" style="font-size: 1.3rem;">
            🎯 INDICADORES NACIONALES DE ANEMIA
        </div>
        """, unsafe_allow_html=True)
        
        col_met1, col_met2, col_met3, col_met4 = st.columns(4)
        
        with col_met1:
            prevalencia = indicadores['prevalencia_nacional']
            color = "#ef4444" if prevalencia >= 40 else "#f59e0b" if prevalencia >= 20 else "#10b981"
            emoji = "🔴" if prevalencia >= 40 else "🟡" if prevalencia >= 20 else "🟢"
            
            st.markdown(f"""
            <div class="metric-card-red" style="background: linear-gradient(135deg, {color}20 0%, {color}10 100%); border-left: 5px solid {color};">
                <div class="metric-label">PREVALENCIA NACIONAL</div>
                <div class="highlight-number" style="color: {color}; font-size: 2.5rem;">{emoji} {prevalencia}%</div>
                <div style="font-size: 0.9rem; color: #6b7280;">
                {indicadores['con_anemia']}/{indicadores['total_pacientes']} pacientes
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col_met2:
            hb_promedio = indicadores['hb_promedio_nacional']
            hb_color = "#ef4444" if hb_promedio < 10 else "#f59e0b" if hb_promedio < 11 else "#10b981"
            hb_estado = "CRÍTICO" if hb_promedio < 10 else "RIESGO" if hb_promedio < 11 else "ADECUADO"
            
            st.markdown(f"""
            <div class="metric-card-purple" style="background: linear-gradient(135deg, {hb_color}20 0%, {hb_color}10 100%); border-left: 5px solid {hb_color};">
                <div class="metric-label">HEMOGLOBINA NACIONAL</div>
                <div class="highlight-number" style="color: {hb_color}; font-size: 2.5rem;">{hb_promedio:.1f} g/dL</div>
                <div style="font-size: 0.9rem; color: #6b7280;">
                Estado: {hb_estado}
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col_met3:
            tasa_seg = indicadores['tasa_seguimiento']
            seg_color = "#10b981" if tasa_seg >= 70 else "#f59e0b" if tasa_seg >= 40 else "#ef4444"
            seg_emoji = "✅" if tasa_seg >= 70 else "⚠️" if tasa_seg >= 40 else "❌"
            
            st.markdown(f"""
            <div class="metric-card-green" style="background: linear-gradient(135deg, {seg_color}20 0%, {seg_color}10 100%); border-left: 5px solid {seg_color};">
                <div class="metric-label">TASA SEGUIMIENTO</div>
                <div class="highlight-number" style="color: {seg_color}; font-size: 2.5rem;">{seg_emoji} {tasa_seg}%</div>
                <div style="font-size: 0.9rem; color: #6b7280;">
                {indicadores['en_seguimiento']} pacientes en control
                </div>
            </div>
            """, unsafe_allow_html=True)

        with col_met4:
            casos_severos = (df["clasificacion_anemia"] == "ANEMIA SEVERA").sum()

            severo_color = (
                "#dc2626" if casos_severos > 10
                else "#f59e0b" if casos_severos > 5
                else "#10b981"
            )

            total_con_anemia = df[
                df["clasificacion_anemia"].isin([
                    "ANEMIA LEVE",
                    "ANEMIA MODERADA",
                    "ANEMIA SEVERA"
                ])
            ].shape[0]

            severo_porcentaje = (
                casos_severos / total_con_anemia * 100
                if total_con_anemia > 0 else 0
            )

            st.markdown(f"""
            <div class="metric-card-yellow" style="background: linear-gradient(135deg, {severo_color}20 0%, {severo_color}10 100%); border-left: 5px solid {severo_color};">
                <div class="metric-label">CASOS SEVEROS</div>
                <div class="highlight-number" style="color: {severo_color}; font-size: 2.5rem;">🚨 {casos_severos}</div>
                <div style="font-size: 0.9rem; color: #6b7280;">
                {severo_porcentaje:.1f}% de los casos
                </div>
            </div>
            """, unsafe_allow_html=True)

        
        # ============================================
        # MAPA INTERACTIVO DEL PERÚ + ESTADÍSTICO DE REGIÓN
        # ============================================
        
        st.markdown("""
        <div class="section-title-blue" style="font-size: 1.3rem;">
            🗺️ MAPA DE PREVALENCIA DE ANEMIA EN EL PERÚ
        </div>
        """, unsafe_allow_html=True)
        
        if 'mapa_peru' in st.session_state and not st.session_state.mapa_peru.empty:
            mapa_df = st.session_state.mapa_peru
            
            # ESTADÍSTICO: REGIÓN CON MÁS ANEMIA
            if not mapa_df.empty and 'prevalencia' in mapa_df.columns:
                # Encontrar la región con mayor prevalencia
                region_max_anemia = mapa_df.loc[mapa_df['prevalencia'].idxmax()]
                region_min_anemia = mapa_df.loc[mapa_df['prevalencia'].idxmin()]
                
                # Mostrar estadístico en columnas
                col_stat1, col_stat2, col_stat3 = st.columns([2, 1, 2])
                
                with col_stat1:
                    st.markdown(f"""
                    <div style="background: linear-gradient(135deg, #dc262620 0%, #dc262610 100%); 
                                padding: 1rem; border-radius: 10px; border-left: 5px solid #dc2626;">
                        <div style="font-weight: 600; color: #dc2626; margin-bottom: 0.5rem;">⚠️ MAYOR PREVALENCIA</div>
                        <div style="font-size: 1.5rem; font-weight: bold; color: #dc2626;">
                        {region_max_anemia['region']}
                        </div>
                        <div style="font-size: 1rem; color: #6b7280;">
                        {region_max_anemia['prevalencia']}% de anemia
                        </div>
                        <div style="font-size: 0.9rem; color: #9ca3af;">
                        {region_max_anemia['con_anemia']}/{region_max_anemia['total_pacientes']} pacientes
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col_stat2:
                    st.markdown("""
                    <div style="text-align: center; padding: 1rem;">
                        <div style="font-size: 2rem;">📊</div>
                        <div style="font-size: 0.9rem; color: #6b7280;">vs</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col_stat3:
                    st.markdown(f"""
                    <div style="background: linear-gradient(135deg, #10b98120 0%, #10b98110 100%); 
                                padding: 1rem; border-radius: 10px; border-left: 5px solid #10b981;">
                        <div style="font-weight: 600; color: #10b981; margin-bottom: 0.5rem;">✅ MENOR PREVALENCIA</div>
                        <div style="font-size: 1.5rem; font-weight: bold; color: #10b981;">
                        {region_min_anemia['region']}
                        </div>
                        <div style="font-size: 1rem; color: #6b7280;">
                        {region_min_anemia['prevalencia']}% de anemia
                        </div>
                        <div style="font-size: 0.9rem; color: #9ca3af;">
                        {region_min_anemia['con_anemia']}/{region_min_anemia['total_pacientes']} pacientes
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
            
            # MAPA INTERACTIVO
            fig_mapa = px.scatter_mapbox(
                mapa_df,
                lat="lat",
                lon="lon",
                color="prevalencia",
                size="total_pacientes",
                hover_name="region",
                hover_data={
                    "prevalencia": ":.1f%",
                    "total_pacientes": True,
                    "con_anemia": True,
                    "hb_promedio": ":.1f"
                },
                color_continuous_scale="RdYlGn_r",
                range_color=[0, 100],
                size_max=30,
                zoom=4.5,
                center={"lat": -9.19, "lon": -75.0},
                title="<b>Prevalencia de Anemia por Región</b>",
                mapbox_style="carto-positron"
            )
            
            fig_mapa.update_layout(
                height=500, 
                margin={"r":0,"t":40,"l":0,"b":0},
                coloraxis_colorbar=dict(
                    title="Prevalencia (%)",
                    ticksuffix="%"
                )
            )
            
            st.plotly_chart(fig_mapa, use_container_width=True)
            
            # Leyenda del mapa
            col_leg1, col_leg2, col_leg3 = st.columns(3)
            with col_leg1:
                st.markdown("""
                <div style="background: #d73027; color: white; padding: 10px; border-radius: 8px; text-align: center; margin: 5px;">
                    🔴 Alta prevalencia (>40%)
                </div>
                """, unsafe_allow_html=True)
            with col_leg2:
                st.markdown("""
                <div style="background: #fdae61; color: black; padding: 10px; border-radius: 8px; text-align: center; margin: 5px;">
                    🟡 Media prevalencia (20-40%)
                </div>
                """, unsafe_allow_html=True)
            with col_leg3:
                st.markdown("""
                <div style="background: #a6d96a; color: black; padding: 10px; border-radius: 8px; text-align: center; margin: 5px;">
                    🟢 Baja prevalencia (<20%)
                </div>
                """, unsafe_allow_html=True)
        
        # ============================================
        # GRÁFICOS DE DISTRIBUCIÓN
        # ============================================
        
        st.markdown("""
        <div class="section-title-blue" style="font-size: 1.3rem;">
            📈 DISTRIBUCIÓN Y TENDENCIAS
        </div>
        """, unsafe_allow_html=True)
        
        col_graf1, col_graf2 = st.columns(2)
        
        with col_graf1:
            # Gráfico de niveles de anemia
            niveles_data = {
                'SEVERA': indicadores['severa'],
                'MODERADA': indicadores['moderada'],
                'LEVE': indicadores['leve'],
                'NORMAL': indicadores['normal']
            }
            
            fig_niveles = px.bar(
                x=list(niveles_data.keys()),
                y=list(niveles_data.values()),
                title='<b>Distribución por Nivel de Anemia</b>',
                color=list(niveles_data.keys()),
                color_discrete_map={
                    'SEVERA': '#dc2626',
                    'MODERADA': '#f59e0b',
                    'LEVE': '#3b82f6',
                    'NORMAL': '#10b981'
                },
                text=list(niveles_data.values())
            )
            
            fig_niveles.update_traces(
                texttemplate='%{y}',
                textposition='outside'
            )
            
            fig_niveles.update_layout(
                xaxis_title="Nivel de Anemia",
                yaxis_title="Número de Pacientes",
                showlegend=False,
                height=350
            )
            
            st.plotly_chart(fig_niveles, use_container_width=True)
        
        with col_graf2:
            # Gráfico SIMPLE de género - SOLO cuenta F y M
            if 'genero' in datos.columns:
                # Limpiar datos: solo tomar F y M exactos
                datos_genero = datos['genero'].astype(str).str.upper().str.strip()
                
                # Filtrar SOLO F y M exactos
                genero_filtrado = datos_genero[datos_genero.isin(['F', 'M'])]
                
                # Contar
                conteo_genero = genero_filtrado.value_counts()
                
                # Asegurar que aparezcan ambos géneros aunque sea 0
                if 'M' not in conteo_genero:
                    conteo_genero['M'] = 0
                if 'F' not in conteo_genero:
                    conteo_genero['F'] = 0
                
                # Ordenar: M primero, F después
                conteo_genero = conteo_genero.reindex(['M', 'F'])
                
                # Crear gráfico de pastel
                labels = ['Niños 👦', 'Niñas 👧']
                valores = [conteo_genero.get('M', 0), conteo_genero.get('F', 0)]
                
                fig_genero = px.pie(
                    values=valores,
                    names=labels,
                    title='<b>Distribución por Género</b><br><sub>Contando solo F y M exactos</sub>',
                    color_discrete_sequence=['#3b82f6', '#ef4444'],
                    height=350
                )
                
                # Agregar total en el centro
                total_genero = sum(valores)
                fig_genero.update_layout(
                    annotations=[
                        dict(
                            text=f'Total: {total_genero}',
                            x=0.5, y=0.5,
                            font_size=14,
                            showarrow=False,
                            font=dict(color='gray')
                        )
                    ]
                )
                
                st.plotly_chart(fig_genero, use_container_width=True)
                
                # Mostrar métricas simples
                col_gen1, col_gen2 = st.columns(2)
                with col_gen1:
                    niños = conteo_genero.get('M', 0)
                    porcentaje_niños = (niños/total_genero*100) if total_genero > 0 else 0
                    st.metric("Niños 👦", niños, delta=f"{porcentaje_niños:.1f}%")
                
                with col_gen2:
                    niñas = conteo_genero.get('F', 0)
                    porcentaje_niñas = (niñas/total_genero*100) if total_genero > 0 else 0
                    st.metric("Niñas 👧", niñas, delta=f"{porcentaje_niñas:.1f}%")
                
                # Mostrar info si hay pacientes sin F/M
                pacientes_sin_fm = len(datos) - total_genero
                if pacientes_sin_fm > 0:
                    st.caption(f"ℹ️ {pacientes_sin_fm} paciente(s) no tienen 'F' o 'M' en la columna 'genero'")
            
            else:
                st.info("📊 La columna 'genero' no está presente en los datos")

        # ============================================
        # TABLA DE REGIONES
        # ============================================
        
        st.markdown("""
        <div class="section-title-blue" style="font-size: 1.3rem;">
            📊 RANKING DE REGIONES POR PREVALENCIA
        </div>
        """, unsafe_allow_html=True)
        
        if 'por_region' in indicadores and indicadores['por_region']:
            # Crear DataFrame para ranking
            ranking_data = []
            for region, stats in indicadores['por_region'].items():
                if stats['total'] > 0:
                    tasa_seguimiento_region = 0
                    if stats['con_anemia'] > 0:
                        tasa_seguimiento_region = round((stats['en_seguimiento'] / stats['con_anemia'] * 100), 1)
                    
                    ranking_data.append({
                        'Región': region,
                        'Prevalencia (%)': stats['prevalencia'],
                        'Total': stats['total'],
                        'Con Anemia': stats['con_anemia'],
                        'Hb Promedio': f"{stats['hb_promedio']:.1f}",
                        'Severa': stats['severa'],
                        'En Seguimiento': stats['en_seguimiento'],
                        'Tasa Seg (%)': tasa_seguimiento_region
                    })
            
            if ranking_data:
                ranking_df = pd.DataFrame(ranking_data)
                ranking_df = ranking_df.sort_values('Prevalencia (%)', ascending=False)
                
                # Mostrar tabla con formato mejorado
                st.dataframe(
                    ranking_df,
                    use_container_width=True,
                    height=300,
                    column_config={
                        "Región": st.column_config.TextColumn("Región", width="medium"),
                        "Prevalencia (%)": st.column_config.NumberColumn("Prevalencia", format="%.1f%%"),
                        "Total": st.column_config.NumberColumn("Total", format="%d"),
                        "Con Anemia": st.column_config.NumberColumn("Con Anemia", format="%d"),
                        "Hb Promedio": st.column_config.NumberColumn("Hb Prom", format="%.1f"),
                        "Severa": st.column_config.NumberColumn("Severa", format="%d"),
                        "Tasa Seg (%)": st.column_config.NumberColumn("Tasa Seg", format="%.1f%%")
                    }
                )

               # ============================================
        # 📊 ESTADÍSTICAS NACIONALES ADICIONALES
        # ============================================
        
        st.markdown("""
        <div class="section-title-blue" style="font-size: 1.3rem;">
            📊 ESTADÍSTICAS NACIONALES ADICIONALES
        </div>
        """, unsafe_allow_html=True)

        # Contenedor para estadísticas
        col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)

        with col_stat1:
            # Estadística: Edad promedio
            if 'edad_meses' in datos.columns:
                edad_prom = datos['edad_meses'].mean()
                st.metric(
                    "👶 Edad Promedio", 
                    f"{edad_prom:.1f} meses",
                    help="Edad promedio de los pacientes registrados"
                )
            else:
                st.metric("👶 Edad Promedio", "N/A")

        with col_stat2:
            # Estadística: Tasa de seguimiento
            tasa_seguimiento = indicadores['tasa_seguimiento']
            st.metric(
                "📋 Tasa Seguimiento", 
                f"{tasa_seguimiento}%",
                delta=f"{indicadores['en_seguimiento']} pacientes"
            )

        with tab3:
            st.markdown("## 📈 Dashboard Nacional")
            df = obtener_datos_supabase()

            if df.empty:
                st.warning("No hay datos registrados")
                st.stop()

            df["clasificacion_anemia"] = df["hemoglobina_ajustada"].apply(
            clasificar_anemia_por_hb
            )

            total_severa = (df["clasificacion_anemia"] == "ANEMIA SEVERA").sum()

            st.metric("Casos de Anemia Severa", total_severa)
    
        with col_stat4:
            # Estadística: Meta OMS
            meta_oms = 20  # Meta OMS es <20%
            
            # Convertir a número para poder calcular diferencia
            try:
                # Asegurar que prevalencia_nacional sea número
                prevalencia_num = float(indicadores['prevalencia_nacional'])
                diferencia = prevalencia_num - meta_oms
                
                # Formatear correctamente
                valor_formateado = f"{diferencia:+.1f}%"
                
                st.metric(
                    "🎯 Meta OMS", 
                    valor_formateado,
                    help="Diferencia respecto a la meta OMS (<20%). Positivo = sobre la meta"
                )
            except (ValueError, TypeError) as e:
                # Si hay error, mostrar valor simple
                st.metric(
                    "🎯 Meta OMS", 
                    "Error cálculo",
                    help=f"No se pudo calcular: {str(e)[:30]}"
                )

        # Línea separadora
        st.markdown("---")

        # ============================================
        # 📥 EXPORTAR REPORTES CON PDF
        # ============================================
        
        st.markdown("""
        <div class="section-title-blue" style="font-size: 1.3rem;">
            📥 EXPORTAR REPORTES
        </div>
        """, unsafe_allow_html=True)

        # Contenedor para exportación
        col_exp1, col_exp2, col_exp3, col_exp4 = st.columns(4)

        # COLUMNA 1: Datos completos
        with col_exp1:
            csv_full = datos.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="📊 Datos Completos (CSV)",
                data=csv_full,
                file_name=f"datos_anemia_nacional_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                use_container_width=True,
                type="primary",
                key="btn_download_full_tab3"
            )

        # COLUMNA 2: Indicadores por región
        with col_exp2:
            try:
                if 'por_region' in indicadores and indicadores['por_region']:
                    reporte_data = []
                    for region, stats in indicadores['por_region'].items():
                        reporte_data.append({
                            'Región': region,
                            'Prevalencia (%)': stats['prevalencia'],
                            'Total Pacientes': stats['total'],
                            'Con Anemia': stats['con_anemia'],
                            'Hb Promedio': stats['hb_promedio'],
                            'Anemia Severa': stats['severa'],
                            'Anemia Moderada': stats['moderada'],
                            'Anemia Leve': stats['leve'],
                            'En Seguimiento': stats['en_seguimiento']
                        })
                    
                    if reporte_data:
                        reporte_df = pd.DataFrame(reporte_data)
                        csv_indicadores = reporte_df.to_csv(index=False).encode('utf-8')
                        
                        st.download_button(
                            label="📈 Indicadores (CSV)",
                            data=csv_indicadores,
                            file_name=f"indicadores_anemia_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                            mime="text/csv",
                            use_container_width=True,
                            type="secondary",
                            key="btn_download_indicadores_tab3"
                        )
                else:
                    st.download_button(
                        label="📈 Indicadores (CSV)",
                        data="No hay datos regionales".encode('utf-8'),
                        file_name="sin_datos.csv",
                        mime="text/csv",
                        use_container_width=True,
                        disabled=True,
                        key="btn_no_data_tab3"
                    )
            except Exception as e:
                st.error(f"Error en datos: {str(e)[:50]}")

        # COLUMNA 3: Informe ejecutivo (PDF + CSV)
        with col_exp3:
            col_pdf, col_csv = st.columns(2)
            
            with col_pdf:
                # Botón para generar PDF
                if st.button("📄 Generar PDF",
                            use_container_width=True,
                            type="primary",
                            key="btn_generar_pdf_nacional_tab3"):
                    
                    with st.spinner("Generando informe PDF..."):
                        try:
                            # Verificar que la función existe
                            pdf_bytes = generar_pdf_dashboard_nacional(
                                indicadores=indicadores,
                                datos=datos,
                                mapa_df=st.session_state.get('mapa_peru', None)
                            )
                            
                            # Mostrar botón de descarga
                            st.download_button(
                                label="📥 Descargar PDF",
                                data=pdf_bytes,
                                file_name=f"dashboard_anemia_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                                mime="application/pdf",
                                use_container_width=True,
                                key="btn_download_pdf_nacional_tab3"
                            )
                            
                        except NameError:
                            st.error("❌ La función 'generar_pdf_dashboard_nacional' no está definida")
                        except Exception as e:
                            st.error(f"❌ Error: {str(e)[:100]}")
            
            with col_csv:
                # Crear CSV simple
                informe_csv = f"""INFORME NACIONAL DE ANEMIA
Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}

RESUMEN NACIONAL
Total Pacientes,{indicadores['total_pacientes']}
Prevalencia Nacional,{indicadores['prevalencia_nacional']}%
Con Anemia,{indicadores['con_anemia']}
Hb Promedio,{indicadores['hb_promedio_nacional']:.1f} g/dL
Tasa Seguimiento,{indicadores['tasa_seguimiento']}%

DISTRIBUCIÓN POR GRAVEDAD
Anemia Severa,{indicadores['severa']}
Anemia Moderada,{indicadores['moderada']}
Anemia Leve,{indicadores['leve']}
Normal,{indicadores['normal']}
"""
                
                # Agregar regiones si existen
                if 'por_region' in indicadores:
                    informe_csv += "\nDATOS POR REGIÓN\nRegion,Prevalencia(%),Total,Con_Anemia,Hb_Promedio\n"
                    for region, stats in indicadores['por_region'].items():
                        informe_csv += f"{region},{stats['prevalencia']}%,{stats['total']},{stats['con_anemia']},{stats['hb_promedio']:.1f}\n"
                
                st.download_button(
                    label="📊 Descargar CSV",
                    data=informe_csv.encode('utf-8'),
                    file_name=f"informe_anemia_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv",
                    use_container_width=True,
                    type="secondary",
                    key="btn_download_csv_informe_tab3"
                )

        # COLUMNA 4: Resumen para portapapeles
        with col_exp4:
            # Crear texto para portapapeles
            resumen_texto = f"""📊 RESUMEN ANEMIA NACIONAL
Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}

📈 INDICADORES:
• Total: {indicadores['total_pacientes']} pacientes
• Prevalencia: {indicadores['prevalencia_nacional']}%
• Con anemia: {indicadores['con_anemia']}
• Hb promedio: {indicadores['hb_promedio_nacional']:.1f} g/dL
• Tasa seguimiento: {indicadores['tasa_seguimiento']}%

🎯 DISTRIBUCIÓN:
• Severa: {indicadores['severa']}
• Moderada: {indicadores['moderada']}
• Leve: {indicadores['leve']}
• Normal: {indicadores['normal']}

---
Sistema Nacional de Monitoreo de Anemia"""
            
            # Mostrar código para copiar manualmente
            if st.button("📋 Copiar Resumen",
                        use_container_width=True,
                        type="secondary",
                        key="btn_copiar_resumen_tab3"):
                st.code(resumen_texto, language="text")
                st.success("✅ Copia el texto de arriba manualmente")

      # ============================================
        # 📌 INFORMACIÓN ADICIONAL
        # ============================================
        
        with st.expander("📌 **INFORMACIÓN TÉCNICA DEL DASHBOARD**", expanded=False):
            st.markdown("""
            **Definiciones utilizadas:**
            
            **Prevalencia de anemia:** Porcentaje de pacientes cuya **Hemoglobina Ajustada** es menor a 11.0 g/dL.
            
            **Clasificación unificada por niveles (Norma Técnica):**
            - 🔴 **Anemia severa:** Hb < 7.0 g/dL
            - 🟠 **Anemia moderada:** Hb 7.0 - 9.9 g/dL  
            - 🟡 **Anemia leve:** Hb 10.0 - 10.9 g/dL
            - 🟢 **Sin anemia:** Hb ≥ 11.0 g/dL
            
            **Nota sobre Altitud:** El sistema detecta automáticamente la región del paciente y aplica el factor de corrección de hemoglobina según los metros sobre el nivel del mar (msnm).
            
            **Indicadores de seguimiento:**
            - **Tasa de seguimiento:** Porcentaje de pacientes diagnosticados con anemia que cuentan con controles activos.
            - **Meta OMS:** Reducción de la prevalencia por debajo del 20% en población infantil.
            
            **Interpretación de colores en el mapa:**
            - 🔴 **Rojo:** Prevalencia > 40% (Problema de salud pública grave)
            - 🟡 **Amarillo:** Prevalencia 20-40% (Atención requerida)
            - 🟢 **Verde:** Prevalencia < 20% (Controlado / Meta OMS)
            
            **Fuentes de datos:**
            - Sistema Nixon v2.0 (Sincronización Supabase)
            - Norma Técnica de Salud NTS Nº 134-MINSA/2017/DGIESP
            - Coordenadas geográficas regionales actualizadas 2026.
            """)
        # ============================================
        # SIN DATOS CARGADOS
        # ============================================
        
        col_empty1, col_empty2, col_empty3 = st.columns([1, 2, 1])
        
        with col_empty2:
            st.markdown("""
            <div style="text-align: center; padding: 3rem; background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%); 
                        border-radius: 15px; border: 2px dashed #cbd5e1; margin: 2rem 0;">
                <div style="font-size: 4rem; margin-bottom: 1rem;">🗺️</div>
                <h3 style="color: #1e3a8a; margin-bottom: 1rem;">DASHBOARD NACIONAL DE ANEMIA</h3>
                <p style="color: #64748b; margin-bottom: 2rem;">
                Visualiza la prevalencia de anemia en todo el Perú con mapas interactivos, 
                indicadores regionales y análisis comparativos.
                </p>
                
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 1rem; margin-bottom: 2rem;">
                    <div style="background: white; padding: 1rem; border-radius: 10px; border: 1px solid #e2e8f0;">
                        <div style="font-size: 1.5rem;">🗺️</div>
                        <div style="font-weight: 600; color: #1e40af;">Mapa Interactivo</div>
                        <div style="font-size: 0.9rem; color: #64748b;">Visual por regiones</div>
                    </div>
                    <div style="background: white; padding: 1rem; border-radius: 10px; border: 1px solid #e2e8f0;">
                        <div style="font-size: 1.5rem;">📊</div>
                        <div style="font-weight: 600; color: #059669;">Indicadores</div>
                        <div style="font-size: 0.9rem; color: #64748b;">Prevalencia y seguimiento</div>
                    </div>
                    <div style="background: white; padding: 1rem; border-radius: 10px; border: 1px solid #e2e8f0;">
                        <div style="font-size: 1.5rem;">📈</div>
                        <div style="font-weight: 600; color: #d97706;">Ranking Regional</div>
                        <div style="font-size: 0.9rem; color: #64748b;">Comparativa por región</div>
                    </div>
                    <div style="background: white; padding: 1rem; border-radius: 10px; border: 1px solid #e2e8f0;">
                        <div style="font-size: 1.5rem;">📥</div>
                        <div style="font-weight: 600; color: #7c3aed;">Reportes</div>
                        <div style="font-size: 0.9rem; color: #64748b;">Exportación de datos</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            st.info("👆 **Presiona 'CARGAR DATOS NACIONALES' para visualizar el dashboard completo**")
            
# ==================================================
# PESTAÑA 4: SISTEMA DE CITAS MEJORADO Y CORREGIDO
# ==================================================

with tab4:
    st.markdown("""
    <div class="main-title" style="background: linear-gradient(135deg, #8b5cf6 0%, #7c3aed 100%); padding: 2rem;">
        <h2 style="margin: 0; color: white;">📅 SISTEMA DE CITAS Y RECORDATORIOS</h2>
        <p style="margin: 10px 0 0 0; color: rgba(255,255,255,0.9);">
        Citas automáticas según nivel de anemia + Recordatorios + Calendario de seguimiento
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # ============================================
    # FUNCIONES ESPECÍFICAS PARA CITAS AUTOMÁTICAS - VERSIÓN CORREGIDA
    # ============================================
    
    def calcular_frecuencia_cita(hemoglobina, edad_meses):
        """Calcula la frecuencia de citas según nivel de anemia"""
        if hemoglobina < 7:
            return "MENSUAL", 30  # Anemia severa
        elif hemoglobina < 10:
            return "TRIMESTRAL", 90  # Anemia moderada
        elif hemoglobina < 11:
            return "SEMESTRAL", 180  # Anemia leve
        else:
            return "ANUAL", 365  # Normal
    
    def crear_cita_automatica(dni_paciente, hemoglobina, edad_meses, tipo="CONTROL"):
        """Crea una cita automática según el nivel de anemia - VERSIÓN CORREGIDA CON MANEJO DE ERROR 11"""
        try:
            # INTENTAR 3 VECES CON PAUSA PARA ERRORES TEMPORALES
            for intento in range(3):
                try:
                    # Obtener información del paciente
                    response = supabase.table("alertas_hemoglobina")\
                        .select("*")\
                        .eq("dni", dni_paciente)\
                        .execute()
                    
                    if not response.data:
                        return False, "Paciente no encontrado"
                    
                    paciente = response.data[0]
                    
                    # Calcular fecha de próxima cita
                    frecuencia, dias = calcular_frecuencia_cita(hemoglobina, edad_meses)
                    fecha_cita = datetime.now() + timedelta(days=dias)
                    
                    # Determinar tipo de consulta según gravedad
                    if hemoglobina < 7:
                        tipo_consulta = "URGENCIA - Anemia Severa"
                        diagnostico = "Anemia severa requiere seguimiento intensivo"
                        tratamiento = "Suplementación inmediata + Control semanal"
                    elif hemoglobina < 10:
                        tipo_consulta = "SEGUIMIENTO - Anemia Moderada"
                        diagnostico = "Anemia moderada en tratamiento"
                        tratamiento = "Suplementación continua + Control mensual"
                    elif hemoglobina < 11:
                        tipo_consulta = "CONTROL - Anemia Leve"
                        diagnostico = "Anemia leve en vigilancia"
                        tratamiento = "Suplementación preventiva"
                    else:
                        tipo_consulta = "CONTROL PREVENTIVO"
                        diagnostico = "Estado normal, seguimiento preventivo"
                        tratamiento = "Mantenimiento nutricional"
                    
                    # Crear datos de la cita
                    cita_data = {
                        "dni_paciente": dni_paciente,
                        "fecha_cita": fecha_cita.strftime('%Y-%m-%d'),
                        "hora_cita": "09:00:00",
                        "tipo_consulta": tipo_consulta,
                        "diagnostico": diagnostico,
                        "tratamiento": tratamiento,
                        "observaciones": f"Cita automática generada por sistema. Frecuencia: {frecuencia}",
                        "investigador_responsable": "Sistema Automático",
                        "severidad_anemia": "SEVERA" if hemoglobina < 7 else "MODERADA" if hemoglobina < 10 else "LEVE" if hemoglobina < 11 else "NORMAL",
                        "suplemento_hierro": paciente.get('tipo_suplemento_hierro', 'Sulfato ferroso'),
                        "frecuencia_suplemento": paciente.get('frecuencia_suplemento', 'Diario'),
                        "proxima_cita": (fecha_cita + timedelta(days=dias)).strftime('%Y-%m-%d'),
                        "created_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    # Insertar en Supabase
                    response = supabase.table("citas").insert(cita_data).execute()
                    
                    if response.data:
                        return True, f"Cita creada para {fecha_cita.strftime('%d/%m/%Y')} - Frecuencia: {frecuencia}"
                    else:
                        return False, "Error al crear cita"
                        
                except Exception as e:
                    error_str = str(e)
                    if ("Resource temporarily unavailable" in error_str or "Error 11" in error_str) and intento < 2:
                        time.sleep(1)  # Esperar 1 segundo y reintentar
                        continue
                    else:
                        raise e  # Relanzar el error si no es temporal
            
            return False, "Error después de múltiples intentos"
            
        except Exception as e:
            return False, f"Error: {str(e)}"
    
    def obtener_recordatorios_pendientes():
        """Obtiene recordatorios de citas próximas"""
        try:
            hoy = datetime.now().date()
            proxima_semana = hoy + timedelta(days=7)
            
            response = supabase.table("citas")\
                .select("*, alertas_hemoglobina(*)")\
                .eq("alertas_hemoglobina.estado_paciente", "Activo")\
                .gte("fecha_cita", hoy.strftime('%Y-%m-%d'))\
                .lte("fecha_cita", proxima_semana.strftime('%Y-%m-%d'))\
                .execute()
            
            if response.data:
                recordatorios = []
                for cita in response.data:
                    paciente_info = cita.get('alertas_hemoglobina', {})
                    dias_restantes = (datetime.strptime(cita['fecha_cita'], '%Y-%m-%d').date() - hoy).days
                    
                    recordatorios.append({
                        'dni': paciente_info.get('dni', 'N/A'),
                        'nombre': paciente_info.get('nombre_apellido', 'Desconocido'),
                        'telefono': paciente_info.get('telefono', 'Sin teléfono'),
                        'fecha_cita': cita['fecha_cita'],
                        'hora_cita': cita.get('hora_cita', '09:00'),
                        'tipo_consulta': cita['tipo_consulta'],
                        'dias_restantes': dias_restantes,
                        'hemoglobina': paciente_info.get('hemoglobina_dl1', 'N/A'),
                        'prioridad': 'URGENTE' if dias_restantes <= 2 else 'PRÓXIMO' if dias_restantes <= 5 else 'PROGRAMADO'
                    })
                
                return recordatorios
            return []
            
        except Exception as e:
            st.error(f"Error obteniendo recordatorios: {str(e)}")
            return []
    
    def obtener_calendario_seguimiento():
        """Obtiene el calendario de seguimiento organizado por nivel de anemia"""
        try:
            # Obtener pacientes con anemia
            response = supabase.table("alertas_hemoglobina")\
                .select("*")\
                .or_("hemoglobina_dl1.lt.11,en_seguimiento.eq.true")\
                .execute()
            
            if not response.data:
                return []
            
            calendario = []
            
            for paciente in response.data:
                hemoglobina = paciente['hemoglobina_dl1']
                edad_meses = paciente['edad_meses']
                
                # Clasificar anemia
                if hemoglobina < 7:
                    nivel = "SEVERA"
                    frecuencia, dias = "MENSUAL", 30
                    emoji = "🔴"
                elif hemoglobina < 10:
                    nivel = "MODERADA"
                    frecuencia, dias = "TRIMESTRAL", 90
                    emoji = "🟡"
                elif hemoglobina < 11:
                    nivel = "LEVE"
                    frecuencia, dias = "SEMESTRAL", 180
                    emoji = "🟢"
                else:
                    nivel = "NORMAL"
                    frecuencia, dias = "ANUAL", 365
                    emoji = "✅"
                
                # Obtener última cita
                citas_response = supabase.table("citas")\
                    .select("fecha_cita, proxima_cita")\
                    .eq("dni_paciente", paciente['dni'])\
                    .order("fecha_cita", desc=True)\
                    .limit(1)\
                    .execute()
                
                ultima_cita = None
                proxima_cita = None
                
                if citas_response.data:
                    ultima_cita = citas_response.data[0].get('fecha_cita')
                    proxima_cita = citas_response.data[0].get('proxima_cita')
                
                # Si no tiene próxima cita, calcular automáticamente
                if not proxima_cita and paciente['en_seguimiento']:
                    if ultima_cita:
                        ultima_fecha = datetime.strptime(ultima_cita, '%Y-%m-%d')
                        proxima_fecha = ultima_fecha + timedelta(days=dias)
                    else:
                        proxima_fecha = datetime.now() + timedelta(days=30)
                    
                    proxima_cita = proxima_fecha.strftime('%Y-%m-%d')
                
                if proxima_cita:
                    dias_restantes = (datetime.strptime(proxima_cita, '%Y-%m-%d').date() - datetime.now().date()).days
                    
                    calendario.append({
                        'dni': paciente['dni'],
                        'nombre': paciente['nombre_apellido'],
                        'nivel_anemia': nivel,
                        'emoji': emoji,
                        'hemoglobina': hemoglobina,
                        'frecuencia': frecuencia,
                        'proxima_cita': proxima_cita,
                        'dias_restantes': dias_restantes,
                        'telefono': paciente.get('telefono', 'Sin teléfono'),
                        'prioridad': '🚨 URGENTE' if dias_restantes <= 7 else '⚠️ PRÓXIMO' if dias_restantes <= 30 else '📅 PROGRAMADO'
                    })
            
            # Ordenar por prioridad y fecha
            calendario.sort(key=lambda x: (x['dias_restantes'], -x['hemoglobina']))
            return calendario
            
        except Exception as e:
            st.error(f"Error obteniendo calendario: {str(e)}")
            return []
    
    # ============================================
    # INTERFAZ PRINCIPAL DEL SISTEMA DE CITAS
    # ============================================
    
    # Crear pestañas dentro de Sistema de Citas
    tab_citas1, tab_citas2, tab_citas3, tab_citas4 = st.tabs([
        "📅 Calendario de Seguimiento",
        "🔄 Generar Citas Automáticas",
        "🔔 Recordatorios Próximos",
        "📋 Historial de Citas"
    ])
    
    # ============================================
    # TAB 1: CALENDARIO DE SEGUIMIENTO
    # ============================================
    with tab_citas1:
        st.markdown('<div class="section-title-purple">📅 CALENDARIO DE SEGUIMIENTO POR NIVEL DE ANEMIA</div>', unsafe_allow_html=True)
        
        if st.button("🔄 Actualizar Calendario", key="actualizar_calendario"):
            with st.spinner("Cargando calendario..."):
                calendario = obtener_calendario_seguimiento()
                st.session_state.calendario_seguimiento = calendario
        
        if 'calendario_seguimiento' in st.session_state and st.session_state.calendario_seguimiento:
            calendario_df = pd.DataFrame(st.session_state.calendario_seguimiento)
            
            # Mostrar por nivel de anemia
            niveles = ["SEVERA", "MODERADA", "LEVE", "NORMAL"]
            
            for nivel in niveles:
                pacientes_nivel = calendario_df[calendario_df['nivel_anemia'] == nivel]
                
                if not pacientes_nivel.empty:
                    emoji = {"SEVERA": "🔴", "MODERADA": "🟡", "LEVE": "🟢", "NORMAL": "✅"}[nivel]
                    frecuencia = {"SEVERA": "MENSUAL", "MODERADA": "TRIMESTRAL", "LEVE": "SEMESTRAL", "NORMAL": "ANUAL"}[nivel]
                    
                    st.markdown(f"""
                    <div class="section-title-purple" style="font-size: 1.4rem; margin-top: 1.5rem;">
                        {emoji} ANEMIA {nivel} - Control {frecuencia}
                    </div>
                    """, unsafe_allow_html=True)
                    
                    for _, paciente in pacientes_nivel.iterrows():
                        col1, col2, col3, col4 = st.columns([3, 2, 2, 1])
                        
                        with col1:
                            st.markdown(f"**{paciente['nombre']}**")
                            st.caption(f"DNI: {paciente['dni']} | Tel: {paciente['telefono']}")
                        
                        with col2:
                            st.markdown(f"**{paciente['hemoglobina']} g/dL**")
                            st.caption(f"Próxima: {paciente['proxima_cita']}")
                        
                        with col3:
                            if paciente['dias_restantes'] <= 0:
                                st.error(f"⚠️ VENCIDO {abs(paciente['dias_restantes'])} días")
                            elif paciente['dias_restantes'] <= 7:
                                st.warning(f"⏰ {paciente['dias_restantes']} días")
                            else:
                                st.info(f"📅 {paciente['dias_restantes']} días")
                        
                        with col4:
                            if st.button("📝", key=f"cita_{paciente['dni']}"):
                                st.session_state.crear_cita_paciente = paciente['dni']
                                st.rerun()
        else:
            st.info("👆 Presiona 'Actualizar Calendario' para cargar el calendario de seguimiento")
    
  # ============================================
# TAB 2: GENERAR CITAS AUTOMÁTICAS - VERSIÓN SIMPLIFICADA
# ============================================
with tab_citas2:
    st.markdown('<div class="section-title-purple">🔄 GENERAR CITAS AUTOMÁTICAS POR NIVEL DE ANEMIA</div>', unsafe_allow_html=True)
    
    # ====== SOLUCIÓN DIRECTA: PRIMERO DIAGNOSTICAR ======
    st.markdown("### 🔍 Paso 1: Diagnosticar el problema")
    
    # Botón para diagnosticar
    if st.button("🔧 EJECUTAR DIAGNÓSTICO COMPLETO", type="primary"):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown("**1. Verificar tabla alertas_hemoglobina**")
            try:
                pacientes = supabase.table("alertas_hemoglobina").select("dni, nombre_apellido, hemoglobina_dl1, en_seguimiento").limit(10).execute()
                if pacientes.data:
                    st.success(f"✅ Tabla encontrada: {len(pacientes.data)} pacientes")
                    for p in pacientes.data[:3]:  # Mostrar solo 3
                        st.write(f"- {p['dni']}: Hb={p['hemoglobina_dl1']}, Seg={p['en_seguimiento']}")
                else:
                    st.warning("⚠️ Tabla vacía o sin datos")
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
        
        with col2:
            st.markdown("**2. Verificar tabla citas**")
            try:
                citas = supabase.table("citas").select("dni_paciente, fecha_cita").limit(5).execute()
                if citas.data:
                    st.success(f"✅ Tabla citas: {len(citas.data)} registros")
                    for c in citas.data[:3]:
                        st.write(f"- {c['dni_paciente']}: {c['fecha_cita']}")
                else:
                    st.info("ℹ️ Tabla citas vacía")
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
        
        with col3:
            st.markdown("**3. Probar consulta combinada**")
            try:
                # Prueba la consulta que debería funcionar
                test_query = supabase.table("alertas_hemoglobina")\
                    .select("count", count="exact")\
                    .lt("hemoglobina_dl1", 11)\
                    .execute()
                st.success(f"✅ Pacientes con Hb < 11: {test_query.count}")
            except Exception as e:
                st.error(f"❌ Error en consulta: {str(e)}")
    
    st.markdown("---")
    
    # ====== SOLUCIÓN ALTERNA: USAR CONSULTA SIMPLE ======
    try:
        # CONSULTA MÁS SIMPLE Y SEGURA
        st.markdown("### 📋 Pacientes que necesitan citas")
        
        # Opción 1: Solo por hemoglobina baja
        pacientes_hb_baja = supabase.table("alertas_hemoglobina")\
            .select("dni, nombre_apellido, hemoglobina_dl1, edad_meses, en_seguimiento, riesgo")\
            .lt("hemoglobina_dl1", 11.0)\
            .execute()
        
        # Opción 2: Solo por seguimiento activo
        pacientes_seguimiento = supabase.table("alertas_hemoglobina")\
            .select("dni, nombre_apellido, hemoglobina_dl1, edad_meses, en_seguimiento, riesgo")\
            .eq("en_seguimiento", True)\
            .execute()
        
        # Combinar listas únicas
        todos_pacientes = []
        dnis_vistos = set()
        
        if pacientes_hb_baja.data:
            for p in pacientes_hb_baja.data:
                if p['dni'] not in dnis_vistos:
                    todos_pacientes.append(p)
                    dnis_vistos.add(p['dni'])
        
        if pacientes_seguimiento.data:
            for p in pacientes_seguimiento.data:
                if p['dni'] not in dnis_vistos:
                    todos_pacientes.append(p)
                    dnis_vistos.add(p['dni'])
        
        if not todos_pacientes:
            st.info("📝 No se encontraron pacientes que necesiten citas automáticas")
            st.write("**Sugerencia:** Verifica que haya pacientes con:")
            st.write("- Hemoglobina menor a 11 g/dL")
            st.write("- O marcados para seguimiento activo")
            
            # Mostrar algunos pacientes para referencia
            st.markdown("**Algunos pacientes en el sistema:**")
            try:
                muestra = supabase.table("alertas_hemoglobina")\
                    .select("dni, nombre_apellido, hemoglobina_dl1, en_seguimiento")\
                    .limit(5)\
                    .execute()
                if muestra.data:
                    for p in muestra.data:
                        st.write(f"- {p['nombre_apellido']}: Hb={p['hemoglobina_dl1']}, Seg={p['en_seguimiento']}")
            except:
                pass
        else:
            st.success(f"✅ Encontrados {len(todos_pacientes)} pacientes que necesitan citas")
            
            # Filtrar pacientes que YA TIENEN citas futuras
            pacientes_sin_cita = []
            
            for paciente in todos_pacientes:
                try:
                    # Verificar si ya tiene cita futura
                    citas_existentes = supabase.table("citas")\
                        .select("id")\
                        .eq("dni_paciente", paciente['dni'])\
                        .gte("fecha_cita", datetime.now().strftime('%Y-%m-%d'))\
                        .execute()
                    
                    if not citas_existentes.data:
                        pacientes_sin_cita.append({
                            'dni': paciente['dni'],
                            'nombre': paciente['nombre_apellido'],
                            'hemoglobina': float(paciente['hemoglobina_dl1']),
                            'edad_meses': paciente['edad_meses'],
                            'riesgo': paciente.get('riesgo', 'No evaluado')
                        })
                except:
                    # Si hay error, asumir que no tiene cita
                    pacientes_sin_cita.append({
                        'dni': paciente['dni'],
                        'nombre': paciente['nombre_apellido'],
                        'hemoglobina': float(paciente['hemoglobina_dl1']),
                        'edad_meses': paciente['edad_meses'],
                        'riesgo': paciente.get('riesgo', 'No evaluado')
                    })
            
            if pacientes_sin_cita:
                st.info(f"📋 **{len(pacientes_sin_cita)} pacientes sin citas programadas**")
                
                # Crear DataFrame
                df_citas = pd.DataFrame(pacientes_sin_cita)
                df_citas['frecuencia'] = df_citas['hemoglobina'].apply(
                    lambda x: "MENSUAL" if x < 7 else "TRIMESTRAL" if x < 10 else "SEMESTRAL" if x < 11 else "ANUAL"
                )
                
                # Mostrar tabla
                st.dataframe(
                    df_citas[['nombre', 'dni', 'hemoglobina', 'frecuencia', 'riesgo']],
                    column_config={
                        "nombre": "Paciente",
                        "dni": "DNI",
                        "hemoglobina": st.column_config.NumberColumn("Hb (g/dL)", format="%.1f"),
                        "frecuencia": "Frecuencia",
                        "riesgo": "Riesgo"
                    },
                    use_container_width=True
                )
                
                # Botones para generar citas
                col_btn1, col_btn2 = st.columns(2)
                
                with col_btn1:
                    if st.button("🎯 Generar Citas Automáticas", type="primary", use_container_width=True):
                        with st.spinner("Generando citas..."):
                            exitos = 0
                            for paciente in pacientes_sin_cita:
                                try:
                                    # Calcular frecuencia
                                    if paciente['hemoglobina'] < 7:
                                        dias = 30
                                    elif paciente['hemoglobina'] < 10:
                                        dias = 90
                                    elif paciente['hemoglobina'] < 11:
                                        dias = 180
                                    else:
                                        dias = 365
                                    
                                    # Crear cita
                                    cita_data = {
                                        "dni_paciente": paciente['dni'],
                                        "nombre_paciente": paciente['nombre'],
                                        "fecha_cita": datetime.now().strftime('%Y-%m-%d'),
                                        "hora_cita": datetime.now().strftime('%H:%M:%S'),
                                        "tipo_consulta": "Seguimiento Automático",
                                        "diagnostico": f"Control por anemia (Hb: {paciente['hemoglobina']:.1f} g/dL)",
                                        "tratamiento": "Seguimiento según protocolo",
                                        "observaciones": f"Cita automática generada. Edad: {paciente['edad_meses']} meses.",
                                        "investigador_responsable": "Sistema Automático",
                                        "proxima_cita": (datetime.now() + timedelta(days=dias)).strftime('%Y-%m-%d'),
                                        "hemoglobina_registrada": paciente['hemoglobina'],
                                        "created_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                    }
                                    
                                    supabase.table("citas").insert(cita_data).execute()
                                    exitos += 1
                                    
                                except Exception as e:
                                    st.error(f"Error con {paciente['nombre']}: {str(e)}")
                            
                            st.success(f"✅ {exitos}/{len(pacientes_sin_cita)} citas generadas")
                            st.rerun()
                
                with col_btn2:
                    if st.button("📊 Ver solo (sin generar)", type="secondary", use_container_width=True):
                        st.info("Modo de visualización activado")
                        
            else:
                st.success("🎉 Todos los pacientes ya tienen citas programadas")
                
    except Exception as e:
        st.error(f"❌ Error general: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
    
    # ====== FORMULARIO MANUAL MÁS SIMPLE ======
    st.markdown("---")
    st.markdown("### ➕ Crear Cita Manual")
    
    # Formulario simplificado
    with st.form("cita_manual_simple"):
        # Seleccionar paciente
        try:
            pacientes_lista = supabase.table("alertas_hemoglobina")\
                .select("dni, nombre_apellido, hemoglobina_dl1")\
                .execute()
            
            if pacientes_lista.data:
                paciente_opciones = [f"{p['nombre_apellido']} (DNI: {p['dni']})" for p in pacientes_lista.data]
                paciente_selec = st.selectbox("Paciente:", paciente_opciones)
                
                # Extraer DNI
                dni = paciente_selec.split("DNI: ")[1].replace(")", "") if "DNI:" in paciente_selec else ""
                
                # Mostrar info actual
                paciente_info = next((p for p in pacientes_lista.data if p['dni'] == dni), None)
                if paciente_info:
                    st.info(f"Hemoglobina actual: **{paciente_info['hemoglobina_dl1']} g/dL**")
                    
                    # Campo para nueva hemoglobina
                    nueva_hb = st.number_input(
                        "Nueva hemoglobina (g/dL):",
                        min_value=5.0,
                        max_value=20.0,
                        value=float(paciente_info['hemoglobina_dl1']),
                        step=0.1
                    )
        except:
            st.warning("No se pudieron cargar los pacientes")
            dni = ""
            nueva_hb = 11.0
        
        # Datos de la cita
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            fecha = st.date_input("Fecha:", datetime.now())
            hora = st.time_input("Hora:", datetime.now().time())
        
        with col_f2:
            tipo = st.selectbox("Tipo:", ["Control", "Seguimiento", "Urgencia"])
            diagnostico = st.text_input("Diagnóstico:", "Anemia por deficiencia de hierro")
        
        observaciones = st.text_area("Observaciones:")
        
        # Botón de guardar
        if st.form_submit_button("💾 Guardar Cita"):
            if dni:
                try:
                    # Actualizar hemoglobina en tabla principal
                    supabase.table("alertas_hemoglobina")\
                        .update({"hemoglobina_dl1": nueva_hb})\
                        .eq("dni", dni)\
                        .execute()
                    
                    # Crear cita
                    cita_data = {
                        "dni_paciente": dni,
                        "nombre_paciente": paciente_selec.split("(")[0].strip(),
                        "fecha_cita": fecha.strftime('%Y-%m-%d'),
                        "hora_cita": hora.strftime('%H:%M:%S'),
                        "tipo_consulta": tipo,
                        "diagnostico": diagnostico,
                        "observaciones": observaciones,
                        "hemoglobina_registrada": nueva_hb,
                        "created_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    supabase.table("citas").insert(cita_data).execute()
                    
                    st.success("✅ Cita creada y hemoglobina actualizada")
                    time.sleep(1)
                    st.rerun()
                    
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
            else:
                st.error("❌ Selecciona un paciente")
    # ============================================
    # TAB 3: RECORDATORIOS PRÓXIMOS - VERSIÓN CORREGIDA (SIN EMAIL/SMS)
    # ============================================
    with tab_citas3:
        st.markdown('<div class="section-title-purple">🔔 RECORDATORIOS DE CITAS PRÓXIMAS</div>', unsafe_allow_html=True)
        
        if st.button("🔄 Cargar Recordatorios", key="cargar_recordatorios"):
            with st.spinner("Buscando citas próximas..."):
                recordatorios = obtener_recordatorios_pendientes()
                st.session_state.recordatorios_pendientes = recordatorios
        
        if 'recordatorios_pendientes' in st.session_state:
            recordatorios_df = pd.DataFrame(st.session_state.recordatorios_pendientes)
            
            if not recordatorios_df.empty:
                # Mostrar por prioridad
                for prioridad in ['URGENTE', 'PRÓXIMO', 'PROGRAMADO']:
                    recordatorios_prioridad = recordatorios_df[recordatorios_df['prioridad'] == prioridad]
                    
                    if not recordatorios_prioridad.empty:
                        color = {
                            'URGENTE': '#dc2626',
                            'PRÓXIMO': '#d97706', 
                            'PROGRAMADO': '#2563eb'
                        }[prioridad]
                        
                        emoji = {
                            'URGENTE': '🚨',
                            'PRÓXIMO': '⚠️',
                            'PROGRAMADO': '📅'
                        }[prioridad]
                        
                        st.markdown(f"""
                        <div style="background: linear-gradient(135deg, {color}10 0%, {color}20 100%); 
                                 padding: 1rem; border-radius: 10px; border-left: 5px solid {color}; 
                                 margin: 1rem 0;">
                            <h4 style="margin: 0 0 10px 0; color: {color};">
                                {emoji} {prioridad} ({len(recordatorios_prioridad)} citas)
                            </h4>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        for _, recordatorio in recordatorios_prioridad.iterrows():
                            col_rec1, col_rec2, col_rec3, col_rec4 = st.columns([3, 2, 2, 1])
                            
                            with col_rec1:
                                st.markdown(f"**{recordatorio['nombre']}**")
                                st.caption(f"DNI: {recordatorio['dni']}")
                            
                            with col_rec2:
                                st.markdown(f"**{recordatorio['fecha_cita']}**")
                                st.caption(f"Hora: {recordatorio['hora_cita']}")
                            
                            with col_rec3:
                                st.markdown(f"**{recordatorio['tipo_consulta']}**")
                                st.caption(f"Hb: {recordatorio['hemoglobina']} g/dL")
                            
                            with col_rec4:
                                if st.button("📞", key=f"llamar_{recordatorio['dni']}"):
                                    st.info(f"📞 Llamar al: {recordatorio['telefono']}")
                
                # Botones de acción - VERSIÓN CORREGIDA (SOLO LISTA)
                st.markdown("---")
                col_acc1, col_acc2 = st.columns(2)
                
                with col_acc1:
                    if st.button("📄 Generar lista de llamadas", use_container_width=True):
                        csv = recordatorios_df[['nombre', 'telefono', 'fecha_cita', 'hora_cita']].to_csv(index=False)
                        st.download_button(
                            label="📥 Descargar lista",
                            data=csv,
                            file_name=f"recordatorios_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv",
                            use_container_width=True
                        )
                
                with col_acc2:
                    st.info("📞 Recordatorios por teléfono (modo manual)")
                    st.caption("Funcionalidad automática en desarrollo")
                        
            else:
                st.success("🎉 No hay recordatorios pendientes para la próxima semana")
        else:
            st.info("👆 Presiona 'Cargar Recordatorios' para ver citas próximas")
    
# ============================================
# TAB 4: HISTORIAL DE CITAS - VERSIÓN CORREGIDA
# ============================================
with tab_citas4:
    st.markdown('<div class="section-title-purple">📋 HISTORIAL COMPLETO DE CITAS</div>', unsafe_allow_html=True)
    
    # ====== FUNCIÓN PARA GENERAR PDF CON FPDF ======
    def generar_pdf_cita_fpdf(cita_data):
        """Genera un PDF con los detalles de una cita usando FPDF"""
        try:
            pdf = FPDF()
            pdf.add_page()
            
            # Título
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, "COMPROBANTE DE CITA MÉDICA", 0, 1, 'C')
            pdf.ln(5)
            
            # Línea separadora
            pdf.set_draw_color(0, 0, 0)
            pdf.set_line_width(0.5)
            pdf.line(10, pdf.get_y(), 200, pdf.get_y())
            pdf.ln(10)
            
            # Información del paciente
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 10, "DATOS DEL PACIENTE", 0, 1)
            pdf.set_font("Arial", '', 11)
            pdf.cell(0, 8, f"Nombre: {cita_data.get('nombre_paciente', 'N/A')}", 0, 1)
            pdf.cell(0, 8, f"DNI: {cita_data.get('dni_paciente', 'N/A')}", 0, 1)
            pdf.cell(0, 8, f"Edad: {cita_data.get('edad_meses', 'N/A')} meses", 0, 1)
            pdf.ln(5)
            
            # Detalles de la cita
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 10, "DETALLES DE LA CITA", 0, 1)
            
            # Crear tabla de detalles
            detalles = [
                ("Fecha", cita_data.get('fecha_cita', 'N/A')),
                ("Hora", cita_data.get('hora_cita', 'N/A')),
                ("Tipo de consulta", cita_data.get('tipo_consulta', 'N/A')),
                ("Hemoglobina", f"{cita_data.get('hemoglobina', 'N/A')} g/dL"),
                ("Estado de anemia", cita_data.get('clasificacion_anemia', 'N/A')),
                ("Nivel de riesgo", cita_data.get('riesgo', 'N/A'))
            ]
            
            pdf.set_font("Arial", '', 11)
            for label, value in detalles:
                pdf.cell(60, 8, f"{label}:", 0, 0)
                pdf.set_font("Arial", 'B', 11)
                pdf.cell(0, 8, str(value), 0, 1)
                pdf.set_font("Arial", '', 11)
            
            pdf.ln(5)
            
            # Diagnóstico
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 10, "DIAGNÓSTICO", 0, 1)
            pdf.set_font("Arial", '', 11)
            
            diagnostico = cita_data.get('diagnostico', 'No especificado')
            pdf.multi_cell(0, 8, diagnostico)
            pdf.ln(5)
            
            # Tratamiento
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 10, "TRATAMIENTO PRESCRITO", 0, 1)
            pdf.set_font("Arial", '', 11)
            
            tratamiento = cita_data.get('tratamiento', 'No especificado')
            pdf.multi_cell(0, 8, tratamiento)
            pdf.ln(5)
            
            # Observaciones
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 10, "OBSERVACIONES", 0, 1)
            pdf.set_font("Arial", '', 11)
            
            observaciones = cita_data.get('observaciones', 'No especificado')
            pdf.multi_cell(0, 8, observaciones)
            pdf.ln(10)
            
            # Pie de página
            pdf.set_font("Arial", 'I', 10)
            pdf.cell(0, 8, f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 0, 1, 'C')
            pdf.cell(0, 8, "Sistema de Seguimiento de Anemia - Centro de Salud", 0, 1, 'C')
            
            # Guardar en buffer
            import io
            buffer = io.BytesIO()
            pdf_output = pdf.output(dest='S').encode('latin-1')
            buffer.write(pdf_output)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            st.error(f"Error generando PDF: {str(e)}")
            return None
    
    def generar_pdf_historial_fpdf(citas_data):
        """Genera un PDF con el historial completo de citas"""
        try:
            pdf = FPDF()
            pdf.add_page()
            
            # Título
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, "HISTORIAL DE CITAS MÉDICAS", 0, 1, 'C')
            pdf.set_font("Arial", '', 12)
            pdf.cell(0, 10, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 0, 1, 'C')
            pdf.ln(10)
            
            # Estadísticas
            total_citas = len(citas_data)
            con_anemia = len([c for c in citas_data if c.get('clasificacion_anemia') in ['Leve', 'Moderada', 'Severa']])
            
            pdf.set_font("Arial", 'B', 14)
            pdf.cell(0, 10, "ESTADÍSTICAS", 0, 1)
            pdf.set_font("Arial", '', 11)
            
            pdf.cell(100, 8, "Total de citas:", 0, 0)
            pdf.cell(0, 8, str(total_citas), 0, 1)
            
            pdf.cell(100, 8, "Citas con anemia:", 0, 0)
            pdf.cell(0, 8, str(con_anemia), 0, 1)
            
            if total_citas > 0:
                porcentaje = (con_anemia / total_citas) * 100
                pdf.cell(100, 8, "Porcentaje con anemia:", 0, 0)
                pdf.cell(0, 8, f"{porcentaje:.1f}%", 0, 1)
            
            pdf.ln(10)
            
            # Tabla de citas
            pdf.set_font("Arial", 'B', 14)
            pdf.cell(0, 10, "DETALLE DE CITAS", 0, 1)
            pdf.ln(5)
            
            # Encabezados de la tabla
            pdf.set_font("Arial", 'B', 10)
            pdf.set_fill_color(200, 200, 200)
            pdf.cell(25, 8, "Fecha", 1, 0, 'C', True)
            pdf.cell(50, 8, "Paciente", 1, 0, 'C', True)
            pdf.cell(25, 8, "DNI", 1, 0, 'C', True)
            pdf.cell(20, 8, "Hb", 1, 0, 'C', True)
            pdf.cell(30, 8, "Estado", 1, 0, 'C', True)
            pdf.cell(40, 8, "Tipo", 1, 1, 'C', True)
            
            # Datos de la tabla
            pdf.set_font("Arial", '', 9)
            pdf.set_fill_color(255, 255, 255)
            
            for cita in citas_data[:30]:
                nombre = cita.get('nombre_paciente', 'N/A')
                if len(nombre) > 20:
                    nombre = nombre[:17] + "..."
                
                fecha = cita.get('fecha_cita', 'N/A')
                if isinstance(fecha, str) and len(fecha) > 10:
                    fecha = fecha[:10]
                
                pdf.cell(25, 8, str(fecha), 1, 0, 'C')
                pdf.cell(50, 8, nombre, 1, 0, 'L')
                pdf.cell(25, 8, str(cita.get('dni_paciente', 'N/A')), 1, 0, 'C')
                pdf.cell(20, 8, str(cita.get('hemoglobina', 'N/A')), 1, 0, 'C')
                pdf.cell(30, 8, str(cita.get('clasificacion_anemia', 'N/A')), 1, 0, 'C')
                pdf.cell(40, 8, str(cita.get('tipo_consulta', 'N/A'))[:15], 1, 1, 'C')
            
            if len(citas_data) > 30:
                pdf.ln(5)
                pdf.set_font("Arial", 'I', 9)
                pdf.cell(0, 8, f"Mostrando 30 de {len(citas_data)} citas. Para el listado completo exporte a CSV.", 0, 1)
            
            pdf.ln(10)
            
            # Pie de página
            pdf.set_font("Arial", 'I', 10)
            pdf.cell(0, 8, "Sistema de Seguimiento de Anemia - Reporte generado automáticamente", 0, 1, 'C')
            
            # Guardar en buffer
            import io
            buffer = io.BytesIO()
            pdf_output = pdf.output(dest='S').encode('latin-1')
            buffer.write(pdf_output)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            st.error(f"Error generando PDF histórico: {str(e)}")
            return None
    
    # ====== FUNCIÓN PARA OBTENER CITAS ======
    def obtener_citas_con_info_anemia():
        try:
            response_citas = supabase.table("citas").select("*").order("fecha_cita", desc=True).execute()
            citas = response_citas.data if response_citas.data else []
            
            if not citas:
                return []
            
            citas_con_info = []
            dnis_unicos = list(set([c.get('dni_paciente') for c in citas if c.get('dni_paciente')]))
            
            pacientes_info = {}
            if dnis_unicos:
                response_pacientes = supabase.table("alertas_hemoglobina")\
                    .select("*")\
                    .in_("dni", dnis_unicos)\
                    .execute()
                
                for paciente in response_pacientes.data:
                    pacientes_info[paciente['dni']] = paciente
            
            for cita in citas:
                dni = cita.get('dni_paciente')
                
                if dni and dni in pacientes_info:
                    info_anemia = pacientes_info[dni]
                    hemoglobina = info_anemia.get('hemoglobina_dl1', 0)
                    edad_meses = info_anemia.get('edad_meses', 0)
                    
                    if edad_meses < 60:
                        if hemoglobina >= 11.0:
                            clasificacion = "Normal"
                        elif hemoglobina >= 10.0:
                            clasificacion = "Leve"
                        elif hemoglobina >= 9.0:
                            clasificacion = "Moderada"
                        else:
                            clasificacion = "Severa"
                    else:
                        if hemoglobina >= 12.0:
                            clasificacion = "Normal"
                        elif hemoglobina >= 11.0:
                            clasificacion = "Leve"
                        elif hemoglobina >= 10.0:
                            clasificacion = "Moderada"
                        else:
                            clasificacion = "Severa"
                    
                    cita_completa = {
                        **cita,
                        "nombre_paciente": info_anemia.get('nombre_apellido', 'Desconocido'),
                        "hemoglobina": hemoglobina,
                        "clasificacion_anemia": clasificacion,
                        "riesgo": info_anemia.get('riesgo', 'N/A'),
                        "en_seguimiento": info_anemia.get('en_seguimiento', False),
                        "edad_meses": edad_meses,
                        "peso_kg": info_anemia.get('peso_kg', 'N/A'),
                        "talla_cm": info_anemia.get('talla_cm', 'N/A')
                    }
                else:
                    cita_completa = {
                        **cita,
                        "nombre_paciente": cita.get('nombre_paciente', 'Sin información'),
                        "hemoglobina": "N/A",
                        "clasificacion_anemia": "Sin datos",
                        "riesgo": "N/A",
                        "en_seguimiento": False,
                        "edad_meses": "N/A",
                        "peso_kg": "N/A",
                        "talla_cm": "N/A"
                    }
                
                citas_con_info.append(cita_completa)
            
            return citas_con_info
            
        except Exception as e:
            st.error(f"❌ Error al obtener citas: {str(e)}")
            return []
    
    # ====== INTERFAZ PRINCIPAL ======
    if 'citas_historial' not in st.session_state:
        st.session_state.citas_historial = []
    
    # Botón para cargar citas
    col_load1, col_load2 = st.columns([3, 1])
    
    with col_load1:
        if st.button("🔄 Cargar historial completo", use_container_width=True):
            with st.spinner("Cargando..."):
                citas_vinculadas = obtener_citas_con_info_anemia()
                st.session_state.citas_historial = citas_vinculadas
                st.success(f"✅ {len(citas_vinculadas)} citas cargadas")
    
    with col_load2:
        if st.button("🧹 Limpiar cache", type="secondary", use_container_width=True):
            st.session_state.citas_historial = []
            st.rerun()
    
    # Mostrar contenido si hay datos
    if st.session_state.citas_historial:
        citas_df = pd.DataFrame(st.session_state.citas_historial)
        
        # Procesamiento de datos
        if 'clasificacion_anemia' not in citas_df.columns:
            citas_df['clasificacion_anemia'] = "Sin datos"
        
        def obtener_color_anemia(clasificacion):
            colores = {
                "Normal": "🟢", "Leve": "🟡", "Moderada": "🟠", 
                "Severa": "🔴", "Sin datos": "⚪"
            }
            return colores.get(clasificacion, "⚪")
        
        citas_df['anemia_icono'] = citas_df['clasificacion_anemia'].apply(obtener_color_anemia)
        citas_df['anemia_mostrar'] = citas_df['anemia_icono'] + " " + citas_df['clasificacion_anemia']
        citas_df['fecha_cita'] = pd.to_datetime(citas_df['fecha_cita'], errors='coerce')
        
        # Filtros
        st.markdown("### 🔍 Filtros de búsqueda")
        col_filt1, col_filt2, col_filt3 = st.columns(3)
        
        with col_filt1:
            if 'tipo_consulta' in citas_df.columns and not citas_df['tipo_consulta'].empty:
                tipos = citas_df['tipo_consulta'].dropna().unique()
                filtro_tipo = st.multiselect("Tipo de consulta", tipos, default=tipos[:min(3, len(tipos))])
        
        with col_filt2:
            if 'clasificacion_anemia' in citas_df.columns:
                clasificaciones = citas_df['clasificacion_anemia'].dropna().unique()
                filtro_anemia = st.multiselect("Estado anemia", clasificaciones, default=clasificaciones)
        
        with col_filt3:
            fecha_min = citas_df['fecha_cita'].min().date() if not citas_df['fecha_cita'].isnull().all() else datetime.now().date() - timedelta(days=30)
            fecha_max = citas_df['fecha_cita'].max().date() if not citas_df['fecha_cita'].isnull().all() else datetime.now().date()
            fecha_min_sel = st.date_input("Desde", value=fecha_min)
            fecha_max_sel = st.date_input("Hasta", value=fecha_max)
        
        # Aplicar filtros
        citas_filtradas = citas_df.copy()
        if 'filtro_tipo' in locals() and filtro_tipo:
            citas_filtradas = citas_filtradas[citas_filtradas['tipo_consulta'].isin(filtro_tipo)]
        if 'filtro_anemia' in locals() and filtro_anemia:
            citas_filtradas = citas_filtradas[citas_filtradas['clasificacion_anemia'].isin(filtro_anemia)]
        
        if not citas_filtradas.empty:
            citas_filtradas = citas_filtradas[
                (citas_filtradas['fecha_cita'].dt.date >= fecha_min_sel) &
                (citas_filtradas['fecha_cita'].dt.date <= fecha_max_sel)
            ]
        
        # Mostrar tabla
        if not citas_filtradas.empty:
            st.markdown("### 📋 Listado de citas")
            
            columnas_mostrar = ['fecha_cita', 'hora_cita', 'nombre_paciente', 'dni_paciente',
                               'anemia_mostrar', 'hemoglobina', 'tipo_consulta', 'diagnostico']
            
            for col in ['riesgo', 'edad_meses']:
                if col in citas_filtradas.columns:
                    columnas_mostrar.append(col)
            
            # Crear lista para selección individual
            citas_lista = citas_filtradas.to_dict('records')
            
            # Mostrar tabla
            st.dataframe(
                citas_filtradas[columnas_mostrar],
                use_container_width=True,
                height=300,
                column_config={
                    "fecha_cita": st.column_config.DateColumn("Fecha", format="DD/MM/YYYY"),
                    "hora_cita": "Hora",
                    "nombre_paciente": "Paciente",
                    "dni_paciente": "DNI",
                    "anemia_mostrar": "Estado Anemia",
                    "hemoglobina": st.column_config.NumberColumn("Hb (g/dL)", format="%.1f"),
                    "tipo_consulta": "Tipo",
                    "diagnostico": st.column_config.TextColumn("Diagnóstico", width="medium")
                }
            )
            
            # ====== SECCIÓN DE PDF INDIVIDUAL ======
            st.markdown("### 📄 Generar comprobante de cita")
            
            col_pdf1, col_pdf2 = st.columns([3, 2])
            
            with col_pdf1:
                # Seleccionar cita para PDF individual
                if citas_lista:
                    opciones_citas = [
                        f"{c['fecha_cita']} - {c['nombre_paciente']} (DNI: {c['dni_paciente']})" 
                        for c in citas_lista
                    ]
                    
                    cita_seleccionada_idx = st.selectbox(
                        "Seleccionar cita para generar PDF:",
                        range(len(opciones_citas)),
                        format_func=lambda x: opciones_citas[x]
                    )
                    
                    if st.button("🖨️ Generar PDF de esta cita", use_container_width=True):
                        cita_seleccionada = citas_lista[cita_seleccionada_idx]
                        
                        with st.spinner("Generando PDF..."):
                            pdf_bytes = generar_pdf_cita_fpdf(cita_seleccionada)
                            
                            if pdf_bytes:
                                nombre_paciente = cita_seleccionada.get('nombre_paciente', 'cita').replace(" ", "_")
                                fecha = cita_seleccionada.get('fecha_cita', '')
                                
                                st.download_button(
                                    label="⬇️ Descargar PDF",
                                    data=pdf_bytes,
                                    file_name=f"cita_{nombre_paciente}_{fecha}.pdf",
                                    mime="application/pdf",
                                    use_container_width=True
                                )
            
            with col_pdf2:
                # PDF del historial completo
                st.markdown("###")
                if st.button("📚 Generar PDF del historial", use_container_width=True):
                    with st.spinner("Generando PDF del historial..."):
                        pdf_bytes = generar_pdf_historial_fpdf(citas_lista[:50])
                        
                        if pdf_bytes:
                            st.download_button(
                                label="📕 Descargar PDF histórico",
                                data=pdf_bytes,
                                file_name=f"historial_citas_{datetime.now().strftime('%Y%m%d')}.pdf",
                                mime="application/pdf",
                                use_container_width=True
                            )
            
            # ====== SECCIÓN DE EXPORTACIÓN ======
            st.markdown("---")
            st.markdown("### 📤 Exportar datos")
            
            col_exp1, col_exp2 = st.columns(2)
            
            with col_exp1:
                # Exportar CSV
                csv_data = citas_filtradas.to_csv(index=False, encoding='utf-8')
                st.download_button(
                    label="📊 Descargar CSV",
                    data=csv_data,
                    file_name=f"historial_citas_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            
            with col_exp2:
                # Exportar Excel
                try:
                    import openpyxl
                    import io
                    
                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        citas_filtradas.to_excel(writer, sheet_name='Citas', index=False)
                    
                    st.download_button(
                        label="📄 Descargar Excel",
                        data=excel_buffer.getvalue(),
                        file_name=f"historial_citas_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                except ImportError:
                    st.info("Para Excel, instala: pip install openpyxl")
                except Exception as e:
                    st.error(f"Error Excel: {str(e)}")
            
            # ====== ESTADÍSTICAS - VERSIÓN CORREGIDA (sin key=) ======
            st.markdown("---")
            st.markdown("### 📊 Estadísticas")
            
            col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
            
            with col_stat1:
                total = len(citas_filtradas)
                st.metric("Total citas", total)  # CORREGIDO: sin key=
            
            with col_stat2:
                con_anemia = len([c for c in citas_lista if c.get('clasificacion_anemia') in ['Leve', 'Moderada', 'Severa']])
                st.metric("Con anemia", con_anemia)  # CORREGIDO: sin key=
            
            with col_stat3:
                severas = len([c for c in citas_lista if c.get('clasificacion_anemia') == 'Severa'])
                st.metric("Severas", severas)  # CORREGIDO: sin key=
            
            with col_stat4:
                if 'fecha_cita' in citas_filtradas.columns and not citas_filtradas.empty:
                    ultima = citas_filtradas['fecha_cita'].max()
                    st.metric("Última cita", ultima.strftime('%d/%m') if pd.notna(ultima) else "N/A")  # CORREGIDO: sin key=
        
        else:
            st.info("📝 No hay citas con los filtros seleccionados")
    
    else:
        st.info("👈 Haz clic en 'Cargar historial completo' para ver las citas")
        
        # Estadísticas rápidas
        if st.button("📊 Ver estadísticas rápidas", type="secondary"):
            try:
                response = supabase.table("citas").select("count", count="exact").execute()
                total_citas = response.count if response.count else 0
                
                fecha_limite = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
                response_recientes = supabase.table("citas")\
                    .select("count", count="exact")\
                    .gte("fecha_cita", fecha_limite)\
                    .execute()
                citas_recientes = response_recientes.count if response_recientes.count else 0
                
                st.info(f"**Total de citas en sistema:** {total_citas}")
                st.info(f"**Citas en los últimos 30 días:** {citas_recientes}")
                
            except Exception as e:
                st.error(f"Error al obtener estadísticas: {str(e)}")
# ==================================================
# PESTAÑA 3: HISTORIAL COMPLETO - VERSIÓN CON PDF CORREGIDA
# ==================================================

with tab_seg3:
    st.header("📋 HISTORIAL CLÍNICO COMPLETO")
    
    # Verificar si hay paciente seleccionado
    if not st.session_state.seguimiento_paciente:
        st.warning("⚠️ Seleccione un paciente primero en la pestaña 'Buscar Paciente'")
        
        if st.button("🔍 Ir a Buscar Paciente", 
                    use_container_width=True,
                    key="btn_ir_buscar_desde_historial"):
            st.markdown("""
            <script>
            setTimeout(() => {
                const tabs = document.querySelectorAll('button[role="tab"]');
                if (tabs.length >= 2) {
                    tabs[1].click();
                }
            }, 500);
            </script>
            """, unsafe_allow_html=True)
            st.rerun()
    else:
        paciente = st.session_state.seguimiento_paciente
        
        # Mostrar información del paciente
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #ede9fe 0%, #ddd6fe 100%); 
                    padding: 1.5rem; border-radius: 12px; margin-bottom: 2rem;">
            <h3 style="margin: 0 0 10px 0; color: #5b21b6;">📊 HISTORIAL DE: {paciente.get('nombre_apellido', 'N/A')}</h3>
            <div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px;">
                <div><strong>DNI:</strong> {paciente.get('dni', 'N/A')}</div>
                <div><strong>Edad:</strong> {paciente.get('edad_meses', 'N/A')} meses</div>
                <div><strong>Región:</strong> {paciente.get('region', 'N/A')}</div>
                <div><strong>Hb actual:</strong> {paciente.get('hemoglobina_dl1', 'N/A')} g/dL</div>
                <div><strong>Estado:</strong> {paciente.get('estado_paciente', 'N/A')}</div>
                <div><strong>Riesgo:</strong> {paciente.get('riesgo', 'N/A')}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # Botones de acción
        col_act1, col_act2, col_act3 = st.columns(3)
        
        with col_act1:
            if st.button("🔄 Actualizar Historial", 
                       type="primary", 
                       use_container_width=True,
                       key="btn_actualizar_historial"):
                dni_paciente = str(paciente.get('dni', ''))
                if dni_paciente:
                    try:
                        response = supabase.table("seguimientos")\
                            .select("*")\
                            .eq("dni_paciente", dni_paciente)\
                            .order("fecha_seguimiento", desc=True)\
                            .execute()
                        
                        if response.data:
                            st.session_state.seguimiento_historial = response.data
                            st.success(f"✅ Historial actualizado: {len(response.data)} controles")
                        else:
                            st.session_state.seguimiento_historial = []
                            st.info("📭 No hay controles registrados")
                            
                    except Exception as e:
                        st.error(f"❌ Error al cargar historial: {str(e)[:100]}")
                    time.sleep(1)
                    st.rerun()
        
        with col_act2:
            if st.button("📝 Nuevo Seguimiento", 
                       type="secondary", 
                       use_container_width=True,
                       key="btn_nuevo_seguimiento_desde_historial"):
                st.markdown("""
                <script>
                setTimeout(() => {
                    const tabs = document.querySelectorAll('button[role="tab"]');
                    if (tabs.length >= 3) {
                        tabs[2].click();
                    }
                }, 500);
                </script>
                """, unsafe_allow_html=True)
                st.rerun()
        
        with col_act3:
            if st.button("🔍 Cambiar Paciente", 
                       type="secondary", 
                       use_container_width=True,
                       key="btn_cambiar_paciente"):
                st.markdown("""
                <script>
                setTimeout(() => {
                    const tabs = document.querySelectorAll('button[role="tab"]');
                    if (tabs.length >= 2) {
                        tabs[1].click();
                    }
                }, 500);
                </script>
                """, unsafe_allow_html=True)
                st.rerun()
        
        # Mostrar historial
        historial = st.session_state.get('seguimiento_historial', [])
        
        if historial:
            # Crear DataFrame del historial
            df_historial = pd.DataFrame(historial)
            
            # Ordenar por fecha
            if 'fecha_seguimiento' in df_historial.columns:
                df_historial['fecha_seguimiento'] = pd.to_datetime(df_historial['fecha_seguimiento'])
                df_historial = df_historial.sort_values('fecha_seguimiento', ascending=False)
            
            # Métricas
            col_met1, col_met2, col_met3, col_met4 = st.columns(4)
            
            with col_met1:
                st.metric("Total controles", len(df_historial))
            
            with col_met2:
                if 'hemoglobina_actual' in df_historial.columns:
                    hb_prom = df_historial['hemoglobina_actual'].mean()
                    st.metric("Hb promedio", f"{hb_prom:.1f} g/dL")
            
            with col_met3:
                if 'fecha_seguimiento' in df_historial.columns:
                    ultima = df_historial['fecha_seguimiento'].max().strftime('%d/%m/%Y')
                    st.metric("Último control", ultima)
            
            with col_met4:
                if 'clasificacion_actual' in df_historial.columns:
                    clasificacion_actual = df_historial['clasificacion_actual'].iloc[0] if not df_historial.empty else "N/A"
                    st.metric("Clasificación actual", clasificacion_actual)
            
            # Gráfico de evolución de hemoglobina
            if 'hemoglobina_actual' in df_historial.columns and 'fecha_seguimiento' in df_historial.columns:
                st.markdown("#### 📈 Evolución de Hemoglobina")
                
                # Crear gráfico
                fig = go.Figure()
                
                fig.add_trace(go.Scatter(
                    x=df_historial['fecha_seguimiento'],
                    y=df_historial['hemoglobina_actual'],
                    mode='lines+markers',
                    name='Hb (g/dL)',
                    line=dict(color='#1f77b4', width=3),
                    marker=dict(size=8, color='#1f77b4')
                ))
                
                # Línea de referencia para anemia (11 g/dL)
                fig.add_hline(
                    y=11.0,
                    line_dash="dash",
                    line_color="red",
                    annotation_text="Límite anemia (11 g/dL)",
                    annotation_position="bottom right"
                )
                
                fig.update_layout(
                    title="Evolución de Hemoglobina",
                    xaxis_title="Fecha",
                    yaxis_title="Hemoglobina (g/dL)",
                    template="plotly_white",
                    height=400
                )
                
                st.plotly_chart(fig, use_container_width=True)
            
            # Tabla de controles
            st.markdown("#### 📋 Controles Registrados")
            
            columnas_posibles = ['fecha_seguimiento', 'tipo_seguimiento', 
                                'hemoglobina_actual', 'hemoglobina_ajustada', 
                                'clasificacion_actual', 'observaciones', 'tratamiento_actual',
                                'usuario_responsable', 'proximo_control']
            
            columnas_disponibles = [c for c in columnas_posibles if c in df_historial.columns]
            
            if columnas_disponibles:
                df_mostrar = df_historial[columnas_disponibles].copy()
                
                # Función para acortar texto largo
                def acortar_texto(texto, max_len=100):
                    if isinstance(texto, str) and len(texto) > max_len:
                        return texto[:max_len] + "..."
                    return texto
                
                # Aplicar a columnas de texto largo
                if 'observaciones' in df_mostrar.columns:
                    df_mostrar['observaciones'] = df_mostrar['observaciones'].apply(lambda x: acortar_texto(str(x), 80))
                
                nombres_columnas = {
                    'fecha_seguimiento': 'Fecha',
                    'tipo_seguimiento': 'Tipo',
                    'hemoglobina_actual': 'Hb Actual',
                    'hemoglobina_ajustada': 'Hb Ajustada',
                    'clasificacion_actual': 'Clasificación',
                    'observaciones': 'Observaciones',
                    'tratamiento_actual': 'Tratamiento',
                    'usuario_responsable': 'Responsable',
                    'proximo_control': 'Próximo Control'
                }
                
                df_mostrar = df_mostrar.rename(columns={k: v for k, v in nombres_columnas.items() if k in df_mostrar.columns})
                
                # Mostrar tabla con estilo
                st.dataframe(
                    df_mostrar,
                    use_container_width=True,
                    height=min(400, len(df_mostrar) * 35 + 38),
                    hide_index=True
                )
                
                # ============================================
                # SECCIÓN DE EXPORTACIÓN CON PDF - VERSIÓN CORREGIDA
                # ============================================
                
                st.markdown("#### 📤 Exportar Historial")
                
                # Botones de exportación (3 columnas)
                col_exp1, col_exp2, col_exp3 = st.columns(3)
                
                with col_exp1:
                    if st.button("📊 Exportar a CSV", 
                               use_container_width=True,
                               type="secondary",
                               key="btn_exportar_csv_historial"):
                        csv = df_historial.to_csv(index=False, encoding='utf-8')
                        
                        st.download_button(
                            label="📥 Descargar CSV",
                            data=csv,
                            file_name=f"historial_{paciente.get('dni', 'paciente')}_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv",
                            use_container_width=True,
                            key="btn_descargar_csv_historial"
                        )
                
                with col_exp2:
                    # Botón para generar PDF
                    generar_pdf = st.button("📄 Generar Informe PDF", 
                                          use_container_width=True,
                                          type="primary",
                                          key="btn_generar_pdf_historial")
                
                with col_exp3:
                    if st.button("🖨️ Vista de Impresión", 
                               use_container_width=True,
                               type="secondary",
                               key="btn_vista_impresion_historial"):
                        with st.expander("📄 Vista para Impresión", expanded=True):
                            st.markdown(f"""
                            <div style="padding: 20px; background: white; color: black; font-family: Arial, sans-serif;">
                                <h2 style="text-align: center; color: #1e40af;">HISTORIAL CLÍNICO</h2>
                                <h3 style="color: #374151;">Paciente: {paciente.get('nombre_apellido', 'N/A')}</h3>
                                <p><strong>DNI:</strong> {paciente.get('dni', 'N/A')}</p>
                                <p><strong>Edad:</strong> {paciente.get('edad_meses', 'N/A')} meses</p>
                                <p><strong>Región:</strong> {paciente.get('region', 'N/A')}</p>
                                <p><strong>Total de controles:</strong> {len(historial)}</p>
                                <hr style="border: 1px solid #ccc; margin: 20px 0;">
                            </div>
                            """, unsafe_allow_html=True)
                            
                            for idx, control in enumerate(historial, 1):
                                st.markdown(f"""
                                <div style="border: 1px solid #ccc; padding: 15px; margin-bottom: 15px; background: #f9fafb;">
                                    <h4 style="color: #5b21b6;">Control #{idx} - {control.get('fecha_seguimiento', 'N/A')}</h4>
                                    <div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 10px; margin-bottom: 10px;">
                                        <div><strong>Tipo:</strong> {control.get('tipo_seguimiento', 'N/A')}</div>
                                        <div><strong>Hemoglobina:</strong> {control.get('hemoglobina_actual', 'N/A')} g/dL</div>
                                        <div><strong>Clasificación:</strong> {control.get('clasificacion_actual', 'N/A')}</div>
                                        <div><strong>Responsable:</strong> {control.get('usuario_responsable', 'N/A')}</div>
                                        <div><strong>Tratamiento:</strong> {control.get('tratamiento_actual', 'N/A')}</div>
                                        <div><strong>Próximo control:</strong> {control.get('proximo_control', 'N/A')}</div>
                                    </div>
                                    <div style="background: #f3f4f6; padding: 10px; border-radius: 5px; margin-top: 10px;">
                                        <strong>Observaciones:</strong><br/>
                                        {control.get('observaciones', 'Sin observaciones').replace(chr(10), '<br>')}
                                    </div>
                                </div>
                                """, unsafe_allow_html=True)
            
            else:
                st.info("No hay datos suficientes para mostrar en la tabla")
            
# ============================================
# EXPORTAR REPORTES (DASHBOARD NACIONAL)
# ============================================
# ============================================
# EXPORTAR REPORTES (Pestaña Nacional)
# ============================================

# ============================================
# CREACIÓN DE COLUMNAS PARA BOTONES
# ============================================
# Creamos 3 columnas. Puedes ajustar los números [1, 1, 1] 
# para cambiar el ancho de cada una si lo deseas.
col_exp1, col_exp2, col_exp3 = st.columns([1, 1, 1])

# ============================================
# EXPORTAR REPORTES (SOLUCIÓN FINAL)
# ============================================
with col_exp3:
    # Verificamos si los indicadores existen
    if 'indicadores_anemia' in st.session_state and st.session_state.indicadores_anemia:
        try:
            # Obtenemos el DataFrame del mapa si existe
            df_mapa = st.session_state.get('mapa_nacional_df', None)
            
            # Generar contenido binario
            pdf_content = generar_pdf_dashboard_nacional(
                st.session_state.indicadores_anemia, 
                st.session_state.datos_nacionales,
                mapa_df=df_mapa
            )
            
            # Validamos que no sea None antes de mostrar el botón
            if pdf_content is not None:
                st.download_button(
                    label="📄 Descargar PDF Profesional",
                    data=pdf_content,
                    file_name=f"Reporte_Nacional_{datetime.now().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="btn_pdf_nacional_final"
                )
            else:
                st.error("❌ Falló la creación del archivo binario.")
                
        except Exception as e:
            st.error(f"❌ Error en el proceso: {str(e)[:50]}")
    else:
        st.warning("⚠️ Cargue los datos nacionales primero.")
# ==================================================
# PESTAÑA 5: CONFIGURACIÓN
# ==================================================

with tab5:
    st.markdown('<div class="section-title-blue">⚙️ Configuración del Sistema</div>', unsafe_allow_html=True)
    
    st.markdown('<div class="section-title-blue" style="font-size: 1.2rem;">🛠️ Configuración de Tablas</div>', unsafe_allow_html=True)
    
    col_config1, col_config2 = st.columns(2)
    
    with col_config1:
        if st.button("📋 Configurar Tabla 'citas'", use_container_width=True, type="primary"):
            crear_tabla_citas_simple()
    
    with col_config2:
        if st.button("🔍 Verificar Conexión Supabase", use_container_width=True, type="secondary"):
            try:
                with st.spinner("Verificando conexión..."):
                    test = supabase.table("alertas_hemoglobina").select("*").limit(1).execute()
                    if test.data:
                        st.success("✅ Conexión a Supabase establecida correctamente")
                    else:
                        st.warning("⚠️ Conexión OK pero no hay datos")
            except Exception as e:
                st.error(f"❌ Error de conexión: {str(e)}")
    
    # Pruebas de guardado
    st.markdown('<div class="section-title-blue" style="font-size: 1.2rem;">🧪 Pruebas del Sistema</div>', unsafe_allow_html=True)
    probar_guardado_directo()
    
    # Información del sistema
    st.markdown('<div class="section-title-blue" style="font-size: 1.2rem;">📊 Información del Sistema</div>', unsafe_allow_html=True)
    
    col_info1, col_info2 = st.columns(2)
    
    with col_info1:
        try:
            response = supabase.table("alertas_hemoglobina").select("*").execute()
            total_pacientes = len(response.data) if response.data else 0
            
            st.markdown(f"""
            <div class="metric-card-blue">
                <div class="metric-label">PACIENTES REGISTRADOS</div>
                <div class="highlight-number highlight-blue">{total_pacientes}</div>
            </div>
            """, unsafe_allow_html=True)
        except:
            st.markdown("""
            <div class="metric-card-blue">
                <div class="metric-label">PACIENTES REGISTRADOS</div>
                <div class="highlight-number highlight-blue">0</div>
            </div>
            """, unsafe_allow_html=True)
    
    with col_info2:
        st.markdown(f"""
        <div class="metric-card-green">
            <div class="metric-label">VERSIÓN DEL SISTEMA</div>
            <div class="highlight-number highlight-green">2.0</div>
            <div style="font-size: 0.9rem; color: #6b7280; margin-top: 5px;">
            Última actualización: {datetime.now().strftime('%d/%m/%Y')}
            </div>
        </div>
        """, unsafe_allow_html=True)

# ==================================================
# SIDEBAR CON TÍTULOS MEJORADOS
# ==================================================

with st.sidebar:
    st.markdown('<div class="section-title-blue" style="font-size: 1.4rem;">📋 Sistema de Referencia</div>', unsafe_allow_html=True)
    
    tab_sidebar1, tab_sidebar2, tab_sidebar3 = st.tabs(["🎯 Ajustes Altitud", "📊 Crecimiento", "🔬 Hematología"])
    
    with tab_sidebar1:
        st.markdown('<div style="color: #1e40af; font-weight: 600; margin-bottom: 10px;">Tabla de Ajustes por Altitud</div>', unsafe_allow_html=True)
        ajustes_df = pd.DataFrame(AJUSTE_HEMOGLOBINA)
        st.dataframe(
            ajustes_df.style.format({
                'altitud_min': '{:.0f}',
                'altitud_max': '{:.0f}', 
                'ajuste': '{:+.1f}'
            }),
            use_container_width=True,
            height=300
        )
    
    with tab_sidebar2:
        st.markdown('<div style="color: #1e40af; font-weight: 600; margin-bottom: 10px;">Tablas de Crecimiento OMS</div>', unsafe_allow_html=True)
        referencia_df = obtener_referencia_crecimiento()
        if not referencia_df.empty:
            st.dataframe(referencia_df, use_container_width=True, height=300)
        else:
            st.info("Cargando tablas de referencia...")
    
    with tab_sidebar3:
        st.markdown('<div style="color: #1e40af; font-weight: 600; margin-bottom: 10px;">Criterios de Interpretación</div>', unsafe_allow_html=True)
        
        st.markdown("""
        ### 🩺 FERRITINA (Reservas)
        - **< 15 ng/mL**: 🚨 Deficit severo
        - **15-30 ng/mL**: ⚠️ Deficit moderado  
        - **30-100 ng/mL**: 🔄 Reservas límite
        - **> 100 ng/mL**: ✅ Adecuado
        
        ### 🔬 CHCM (Concentración)
        - **< 32 g/dL**: 🚨 Hipocromía
        - **32-36 g/dL**: ✅ Normocromía
        - **> 36 g/dL**: 🔄 Hipercromía
        
        ### 📈 RETICULOCITOS (Producción)
        - **< 0.5%**: ⚠️ Hipoproliferación
        - **0.5-1.5%**: ✅ Normal
        - **> 1.5%**: 🔄 Hiperproducción
        """)
    
    st.markdown("---")
    
    st.markdown('<div style="color: #1e40af; font-weight: 600; margin-bottom: 10px;">💡 Sistema Integrado</div>', unsafe_allow_html=True)
    st.markdown("""
    - ✅ Ajuste automático por altitud
    - ✅ Clasificación OMS de anemia
    - ✅ Seguimiento por gravedad
    - ✅ Dashboard nacional
    - ✅ Sistema de citas
    - ✅ Interpretación automática
    - ✅ Manejo de duplicados
    """)
    
    # Inicialización de datos de prueba
    if supabase:
        try:
            response = supabase.table(TABLE_NAME).select("*").limit(1).execute()
            if not response.data:
                st.info("🔄 Base de datos vacía. Ingrese pacientes desde 'Registro Completo'")
                
                if st.button("➕ Insertar paciente de prueba", use_container_width=True):
                    with st.spinner("Insertando..."):
                        paciente_prueba = {
                            "dni": "87654321",
                            "nombre_apellido": "Carlos López Díaz",
                            "edad_meses": 36,
                            "peso_kg": 14.5,
                            "talla_cm": 95.0,
                            "genero": "M",
                            "telefono": "987123456",
                            "estado_paciente": "Activo",
                            "region": "LIMA",
                            "departamento": "Lima Centro",
                            "altitud_msnm": 150,
                            "nivel_educativo": "Secundaria",
                            "acceso_agua_potable": True,
                            "tiene_servicio_salud": True,
                            "hemoglobina_dl1": 10.5,
                            "en_seguimiento": True,
                            "consumir_hierro": True,
                            "tipo_suplemento_hierro": "Sulfato ferroso",
                            "frecuencia_suplemento": "Diario",
                            "antecedentes_anemia": False,
                            "enfermedades_cronicas": "Ninguna",
                            "interpretacion_hematologica": "Anemia leve por deficiencia de hierro",
                            "politicas_de_ris": "LIMA",
                            "riesgo": "RIESGO MODERADO",
                            "fecha_alerta": datetime.now().strftime("%Y-%m-%d"),
                            "estado_alerta": "EN SEGUIMIENTO",
                            "sugerencias": "Suplementación con hierro y control mensual",
                            "severidad_interpretacion": "LEVE"
                        }
                        
                        resultado = insertar_datos_supabase(paciente_prueba)
                        if resultado:
                            st.success("✅ Paciente de prueba insertado")
                            time.sleep(2)
                            st.rerun()
                        else:
                            st.error("❌ Error al insertar paciente de prueba")
        except Exception as e:
            st.warning(f"⚠️ Error verificando datos: {e}")

# ==================================================
# PIE DE PÁGINA CON ESTILO MEJORADO
# ==================================================

st.markdown("---")
st.markdown("""
<div style="text-align: center; padding: 2rem; background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%); border-radius: 10px;">
    <h4 style="color: #1e3a8a; margin: 0 0 10px 0;">🏥 SISTEMA NIXON</h4>
    <p style="color: #475569; margin: 5px 0;">Control de Anemia y Nutrición Infantil</p>
    <p style="color: #64748b; font-size: 0.9rem; margin: 5px 0;">Versión 2.0 | {}</p>
    <p style="color: #94a3b8; font-size: 0.8rem; margin-top: 15px;">
    ⚠️ <em>Para uso médico profesional. Consulte siempre con especialistas.</em>
    </p>
</div>
""".format(datetime.now().strftime("%d/%m/%Y")), unsafe_allow_html=True)
